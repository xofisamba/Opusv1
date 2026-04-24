"""CSV/Excel export utilities for waterfall results.

Provides functions to export:
- Waterfall period data (CSV)
- Summary metrics (CSV)
- Financial statements (CSV)

Usage:
    from utils.export import export_waterfall_csv
    export_waterfall_csv(result, "waterfall_output.csv")
"""
import csv
from pathlib import Path
from typing import Optional

from domain.waterfall.waterfall_engine import WaterfallResult


def export_waterfall_csv(result: WaterfallResult, filepath: str) -> None:
    """Export waterfall period data to CSV.
    
    Args:
        result: WaterfallResult from run_waterfall
        filepath: Output CSV file path
    """
    fieldnames = [
        "period", "year_index", "period_in_year", "is_operation",
        "generation_mwh", "revenue_keur", "opex_keur", "ebitda_keur",
        "depreciation_keur", "interest_senior_keur", "interest_shl_keur",
        "tax_keur", "cf_after_tax_keur",
        "senior_ds_keur", "senior_interest_keur", "senior_principal_keur",
        "shl_service_keur", "shl_interest_keur", "shl_principal_keur",
        "dsra_contribution_keur", "dsra_balance_keur",
        "cf_after_reserves_keur", "dscr", "llcr", "plcr",
        "lockup_active", "distribution_keur", "cash_sweep_keur",
        "cum_distribution_keur", "cash_balance_keur",
    ]
    
    with open(filepath, "w", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        
        for p in result.periods:
            writer.writerow({
                "period": p.period,
                "year_index": p.year_index,
                "period_in_year": p.period_in_year,
                "is_operation": p.is_operation,
                "generation_mwh": round(p.generation_mwh, 2),
                "revenue_keur": round(p.revenue_keur, 2),
                "opex_keur": round(p.opex_keur, 2),
                "ebitda_keur": round(p.ebitda_keur, 2),
                "depreciation_keur": round(p.depreciation_keur, 2),
                "interest_senior_keur": round(p.interest_senior_keur, 2),
                "interest_shl_keur": round(p.interest_shl_keur, 2),
                "tax_keur": round(p.tax_keur, 2),
                "cf_after_tax_keur": round(p.cf_after_tax_keur, 2),
                "senior_ds_keur": round(p.senior_ds_keur, 2),
                "senior_interest_keur": round(p.senior_interest_keur, 2),
                "senior_principal_keur": round(p.senior_principal_keur, 2),
                "shl_service_keur": round(p.shl_service_keur, 2),
                "shl_interest_keur": round(p.shl_interest_keur, 2),
                "shl_principal_keur": round(p.shl_principal_keur, 2),
                "dsra_contribution_keur": round(p.dsra_contribution_keur, 2),
                "dsra_balance_keur": round(p.dsra_balance_keur, 2),
                "cf_after_reserves_keur": round(p.cf_after_reserves_keur, 2),
                "dscr": round(p.dscr, 4) if p.dscr < float('inf') else 999.0,
                "llcr": round(p.llcr, 4) if p.llcr < float('inf') else 999.0,
                "plcr": round(p.plcr, 4) if p.plcr < float('inf') else 999.0,
                "lockup_active": p.lockup_active,
                "distribution_keur": round(p.distribution_keur, 2),
                "cash_sweep_keur": round(p.cash_sweep_keur, 2),
                "cum_distribution_keur": round(p.cum_distribution_keur, 2),
                "cash_balance_keur": round(p.cash_balance_keur, 2),
            })


def export_summary_csv(result: WaterfallResult, filepath: str) -> None:
    """Export summary metrics to CSV.
    
    Args:
        result: WaterfallResult from run_waterfall
        filepath: Output CSV file path
    """
    sculpt = result.sculpting_result
    
    rows = [
        ("=== REVENUE ===", ""),
        ("Total Revenue (kEUR)", f"{result.total_revenue_keur:,.0f}"),
        ("Total OPEX (kEUR)", f"{result.total_opex_keur:,.0f}"),
        ("Total EBITDA (kEUR)", f"{result.total_ebitda_keur:,.0f}"),
        ("Total Tax (kEUR)", f"{result.total_tax_keur:,.0f}"),
        ("", ""),
        ("=== DEBT ===", ""),
        ("Debt Amount (kEUR)", f"{sculpt.debt_keur:,.0f}"),
        ("Total Senior DS (kEUR)", f"{result.total_senior_ds_keur:,.0f}"),
        ("Total SHL Service (kEUR)", f"{result.total_shl_service_keur:,.0f}"),
        ("", ""),
        ("=== RETURNS ===", ""),
        ("Project IRR", f"{result.project_irr*100:.3f}%"),
        ("Equity IRR", f"{result.equity_irr*100:.3f}%" if result.equity_irr else "N/A"),
        ("Project NPV (kEUR)", f"{result.project_npv:,.0f}"),
        ("Equity NPV (kEUR)", f"{result.equity_npv:,.0f}" if result.equity_npv else "N/A"),
        ("", ""),
        ("=== COVENANTS ===", ""),
        ("Avg DSCR", f"{result.avg_dscr:.3f}"),
        ("Min DSCR", f"{result.min_dscr:.3f}"),
        ("Min LLCR", f"{result.min_llcr:.3f}" if result.min_llcr else "N/A"),
        ("Min PLCR", f"{result.min_plcr:.3f}" if result.min_plcr else "N/A"),
        ("Periods in Lockup", f"{result.periods_in_lockup}"),
        ("", ""),
        ("=== DISTRIBUTIONS ===", ""),
        ("Total Distribution (kEUR)", f"{result.total_distribution_keur:,.0f}"),
    ]
    
    with open(filepath, "w", newline="") as f:
        writer = csv.writer(f)
        for row in rows:
            writer.writerow(row)


def waterfall_to_dataframe(result: WaterfallResult) -> "pd.DataFrame":
    """Convert waterfall result to pandas DataFrame.
    
    Args:
        result: WaterfallResult from run_waterfall
    
    Returns:
        DataFrame with period-level data
    """
    import pandas as pd
    
    data = []
    for p in result.periods:
        data.append({
            "Period": p.period,
            "Year": p.year_index,
            "H": p.period_in_year,
            "Op": p.is_operation,
            "Gen (MWh)": round(p.generation_mwh, 0),
            "Rev (kEUR)": round(p.revenue_keur, 0),
            "OPEX (kEUR)": round(p.opex_keur, 0),
            "EBITDA (kEUR)": round(p.ebitda_keur, 0),
            "Dep (kEUR)": round(p.depreciation_keur, 0),
            "Int Sen (kEUR)": round(p.interest_senior_keur, 0),
            "Tax (kEUR)": round(p.tax_keur, 0),
            "CFAT (kEUR)": round(p.cf_after_tax_keur, 0),
            "Sen DS (kEUR)": round(p.senior_ds_keur, 0),
            "DSCR": round(p.dscr, 2) if p.dscr < float('inf') else None,
            "Dist (kEUR)": round(p.distribution_keur, 0),
            "Sweep (kEUR)": round(p.cash_sweep_keur, 0),
            "Cash Bal (kEUR)": round(p.cash_balance_keur, 0),
        })
    
    return pd.DataFrame(data)


def export_waterfall_excel(result: "WaterfallResult", filepath: str) -> None:
    """Export complete waterfall analysis to formatted Excel workbook.
    
    Creates a multi-sheet workbook suitable for bank/investor presentation:
    - Sheet 1: Summary (key metrics)
    - Sheet 2: Waterfall (period-level cash flows)
    - Sheet 3: Debt Schedule
    - Sheet 4: Covenant Compliance
    
    Args:
        result: WaterfallResult from run_waterfall
        filepath: Output Excel file path
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    
    wb = Workbook()
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4472C4")
    number_format = "#,##0"
    pct_format = "0.00%"
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    def style_header(ws, row=1):
        for cell in ws[row]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border
    
    def set_col_widths(ws, widths):
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
    
    # ===== SHEET 1: Summary =====
    ws = wb.active
    ws.title = "Summary"
    
    sc = result.sculpting_result
    summary_data = [
        ("Project Finance Summary", ""),
        ("", ""),
        ("Debt Sizing", ""),
        ("Debt (kEUR)", f"{sc.debt_keur:,.0f}"),
        ("Avg DSCR", f"{result.avg_dscr:.3f}"),
        ("Min DSCR", f"{result.min_dscr:.3f}"),
        ("", ""),
        ("Returns", ""),
        ("Project IRR", f"{result.project_irr*100:.2f}%"),
        ("Equity IRR", f"{result.equity_irr*100:.2f}%" if result.equity_irr else "N/A"),
        ("", ""),
        ("Cash Flows (kEUR)", ""),
        ("Total Revenue", f"{result.total_revenue_keur:,.0f}"),
        ("Total Distributions", f"{result.total_distribution_keur:,.0f}"),
        ("Total Senior Debt Service", f"{result.total_senior_ds_keur:,.0f}"),
        ("Total Tax", f"{result.total_tax_keur:,.0f}"),
    ]
    
    for row_idx, (label, value) in enumerate(summary_data, 1):
        ws.cell(row=row_idx, column=1, value=label)
        ws.cell(row=row_idx, column=2, value=value)
        if label and not value:
            ws.cell(row=row_idx, column=1).font = Font(bold=True)
    
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 20
    
    # ===== SHEET 2: Waterfall =====
    ws2 = wb.create_sheet("Waterfall")
    
    wf_headers = ["Period", "Year", "Half", "Gen (MWh)", "Rev (kEUR)", "EBITDA (kEUR)", 
                  "CFAT (kEUR)", "Sen DS (kEUR)", "DSCR", "Dist (kEUR)", "Sweep (kEUR)", "Lockup"]
    wf_data = []
    for p in result.periods:
        wf_data.append([
            p.period, p.year_index, "H1" if p.period_in_year == 1 else "H2",
            round(p.generation_mwh, 0), round(p.revenue_keur, 0), round(p.ebitda_keur, 0),
            round(p.cf_after_tax_keur, 0), round(p.senior_ds_keur, 0),
            round(p.dscr, 3) if p.dscr < float('inf') else 999,
            round(p.distribution_keur, 0), round(p.cash_sweep_keur, 0),
            "Y" if p.lockup_active else "N"
        ])
    
    for col, header in enumerate(wf_headers, 1):
        ws2.cell(row=1, column=col, value=header)
    style_header(ws2)
    for row_idx, row_data in enumerate(wf_data, 2):
        for col_idx, val in enumerate(row_data, 1):
            cell = ws2.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border
    
    set_col_widths(ws2, [8, 6, 6, 12, 12, 12, 12, 12, 8, 12, 12, 8])
    
    # ===== SHEET 3: Debt Schedule =====
    ws3 = wb.create_sheet("Debt Schedule")
    
    ds_headers = ["Period", "Year", "Opening Bal", "Interest", "Principal", "Closing Bal", "DSCR"]
    ds_data = []
    for i, (bal, ir, pr) in enumerate(zip(sc.balance_schedule, sc.interest_schedule, sc.principal_schedule)):
        period = i
        year = i // 2 + 1
        half = "H1" if i % 2 == 0 else "H2"
        ds_data.append([period, year, half, round(bal, 0), round(ir, 0), round(pr, 0), round(sc.balance_schedule[i] if i < len(sc.balance_schedule) else 0, 0), round(sc.dscr_schedule[i], 3) if i < len(sc.dscr_schedule) and sc.dscr_schedule[i] < float('inf') else 999])
    
    for col, header in enumerate(ds_headers, 1):
        ws3.cell(row=1, column=col, value=header)
    style_header(ws3)
    for row_idx, row_data in enumerate(ds_data, 2):
        for col_idx, val in enumerate(row_data, 1):
            cell = ws3.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border
    
    set_col_widths(ws3, [8, 6, 6, 14, 14, 14, 8])
    
    # ===== SHEET 4: Covenant =====
    ws4 = wb.create_sheet("Covenant")
    
    cov_headers = ["Year", "DSCR", "LLCR", "PLCR", "Lockup", "DSCR OK", "LLCR OK", "PLCR OK"]
    cov_data = []
    for p in result.periods:
        if p.is_operation and p.period_in_year == 2:
            dscr = p.dscr if p.dscr < float('inf') else 999
            llcr = p.llcr if p.llcr < float('inf') else 999
            plcr = p.plcr if p.plcr < float('inf') else 999
            cov_data.append([
                p.year_index, round(dscr, 3), round(llcr, 3), round(plcr, 3),
                "Y" if p.lockup_active else "N",
                "OK" if dscr >= 1.15 else ("WARN" if dscr >= 1.10 else "BREACH"),
                "OK" if llcr >= 1.15 else "BREACH",
                "OK" if plcr >= 1.20 else "BREACH",
            ])
    
    for col, header in enumerate(cov_headers, 1):
        ws4.cell(row=1, column=col, value=header)
    style_header(ws4)
    for row_idx, row_data in enumerate(cov_data, 2):
        for col_idx, val in enumerate(row_data, 1):
            cell = ws4.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border
            # Color code status columns
            if col_idx >= 6 and isinstance(val, str):
                if val == "BREACH":
                    cell.fill = PatternFill("solid", fgColor="FF6B6B")
                    cell.font = Font(bold=True, color="FFFFFF")
                elif val == "WARN":
                    cell.fill = PatternFill("solid", fgColor="FFE066")
    
    set_col_widths(ws4, [8, 8, 8, 8, 8, 10, 10, 10])
    
    wb.save(filepath)
    print(f"Excel export saved to: {filepath}")
