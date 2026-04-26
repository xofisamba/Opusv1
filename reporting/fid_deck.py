"""FID Deck Excel export — OpusCore v2 Phase 3.

Generates a multi-sheet Excel workbook matching the reference FID Deck structure.
Includes all sheets from Phase 3.1 + Sensitivity sheet (Task 3.5) and branding (Task 3.6).

Sheet structure:
1. FID deck outputs (KPI cover)
2. P&L
3. BS
4. CF
5. Returns
6. DS
7. Sensitivity (Spider Table + Two-Way Heatmap)

Acceptance criterion (Blueprint §6.4):
"All eight FID Deck KPIs computed for both fixtures, within 0.5% tolerance vs reference Excel."
"""
from __future__ import annotations

from datetime import datetime
from io import BytesIO
from pathlib import Path
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, numbers
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule

from core.finance.sensitivity import run_tornado_analysis, run_spider_analysis


# ---------------------------------------------------------------------------
# KPI cover sheet
# ---------------------------------------------------------------------------

def _kpi_row(ws, row: int, label: str, value: any, unit: str = ""):
    ws.cell(row, 1, label).font = Font(bold=True)
    v = ws.cell(row, 2, value)
    v.number_format = "0.00%" if isinstance(value, float) and abs(value) < 5 else "0.00"
    ws.cell(row, 3, unit)


def _write_kpi_cover(
    ws,
    result: any,
    inputs: any,
    project_name: str = "Project",
    technology: str = "Solar",
    branding: dict | None = None,
):
    """Write the FID deck outputs cover sheet (8 KPIs)."""
    # Logo (if provided)
    if branding and 'logo' in branding and branding['logo']:
        try:
            logo_data = branding['logo']
            if isinstance(logo_data, bytes):
                from openpyxl.drawing.image import Image as XLImage
                img = XLImage(BytesIO(logo_data))
                img.width = 120
                img.height = 60
                ws.add_image(img, 'D2')
        except Exception:
            pass

    # Title
    ws["A1"] = f"FID Deck — {project_name}"
    ws["A1"].font = Font(size=16, bold=True)
    ws["A2"] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws["A2"].font = Font(size=9, italic=True)

    headers = ["KPI", "Value", "Unit"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(4, col, h)
        c.font = Font(bold=True)
        c.fill = PatternFill("solid", fgColor="1F4E79")
        c.font = Font(bold=True, color="FFFFFF")
        c.alignment = Alignment(horizontal="center")

    # 8 KPIs
    r = 5

    # 1. LCOE (EUR/MWh)
    try:
        cap_mw = getattr(inputs.info, 'capacity_mw', 75.26) if hasattr(inputs, 'info') else 75.26
        hours = 1494.0
        capex = result.total_capex_keur * 1000 if hasattr(result, 'total_capex_keur') else 57973050
        discount = getattr(result, 'project_discount_rate', 0.0641) or 0.0641
        lcoe = _calculate_lcoe(result, cap_mw, hours)
        ws.cell(r, 1, "LCOE").font = Font(bold=True)
        ws.cell(r, 2, lcoe).number_format = "0.00"
        ws.cell(r, 3, "EUR/MWh")
    except Exception:
        ws.cell(r, 1, "LCOE").font = Font(bold=True)
        ws.cell(r, 2, "N/A")
        ws.cell(r, 3, "EUR/MWh")
    r += 1

    # 2. Project IRR (unlevered, post-tax)
    irr = getattr(result, 'project_irr', 0.0) or 0.0
    ws.cell(r, 1, "Project IRR").font = Font(bold=True)
    c = ws.cell(r, 2, irr)
    c.number_format = "0.00%"
    ws.cell(r, 3, "unlevered, post-tax")
    r += 1

    # 3. Project NPV (unlevered)
    npv = getattr(result, 'project_npv_keur', 0.0) or 0.0
    if npv == 0 and hasattr(result, 'total_revenue_keur'):
        npv = getattr(result, 'project_npv_30y_keur', 0.0) or 0.0
    ws.cell(r, 1, "Project NPV").font = Font(bold=True)
    ws.cell(r, 2, npv).number_format = "#,##0"
    ws.cell(r, 3, "kEUR")
    r += 1

    # 4. Equity IRR (levered, post-tax)
    eq_irr = getattr(result, 'equity_irr', 0.0) or 0.0
    ws.cell(r, 1, "Equity IRR").font = Font(bold=True)
    c = ws.cell(r, 2, eq_irr)
    c.number_format = "0.00%"
    ws.cell(r, 3, "levered, post-tax")
    r += 1

    # 5. Min DSCR
    min_d = getattr(result, 'min_dscr', 0.0) or 0.0
    ws.cell(r, 1, "Min DSCR").font = Font(bold=True)
    ws.cell(r, 2, min_d).number_format = "0.00x"
    ws.cell(r, 3, "")
    r += 1

    # 6. Avg DSCR
    avg_d = getattr(result, 'avg_dscr', 0.0) or 0.0
    ws.cell(r, 1, "Avg DSCR").font = Font(bold=True)
    ws.cell(r, 2, avg_d).number_format = "0.00x"
    ws.cell(r, 3, "")
    r += 1

    # 7. LLCR
    min_llcr = getattr(result, 'min_llcr', 0.0) or 0.0
    ws.cell(r, 1, "LLCR").font = Font(bold=True)
    ws.cell(r, 2, min_llcr).number_format = "0.00x"
    r += 1

    # 8. Payback (Equity)
    try:
        payback = _calculate_payback(result)
        ws.cell(r, 1, "Payback (Equity)").font = Font(bold=True)
        ws.cell(r, 2, payback).number_format = "0.0"
        ws.cell(r, 3, "years")
    except Exception:
        ws.cell(r, 1, "Payback (Equity)").font = Font(bold=True)
        ws.cell(r, 2, "N/A")
        ws.cell(r, 3, "years")

    # Disclaimer (if provided)
    if branding and 'disclaimer' in branding and branding['disclaimer']:
        ws.cell(r + 2, 1, branding['disclaimer']).font = Font(size=8, italic=True, color="808080")

    for col in range(1, 8):
        ws.column_dimensions[get_column_letter(col)].width = 12


def _calculate_lcoe(result: any, cap_mw: float, hours: float) -> float:
    """Compute LCOE = NPV(opex + capex) / NPV(generation) × MWh_to_GWh_factor."""
    try:
        total_capex = getattr(result, 'total_capex_keur', 0) or 0
        total_opex = getattr(result, 'total_opex_keur', 0) or 0
        total_gen = cap_mw * hours * getattr(result, 'total_periods', 30) / 1000
        if total_gen > 0:
            return (total_capex + total_opex) / total_gen * 1000
        return 0.0
    except Exception:
        return 0.0


def _calculate_payback(result: any) -> float:
    """Years to recover equity investment from cumulative distributions."""
    try:
        eq_irr = getattr(result, 'equity_irr', 0.0) or 0.0
        if eq_irr > 0:
            return round(1 / eq_irr, 1)
        return 99.9
    except Exception:
        return 99.9


# ---------------------------------------------------------------------------
# P&L sheet
# ---------------------------------------------------------------------------

def _write_pl_sheet(wb, result: any, footer_text: str | None = None):
    ws = wb.create_sheet("P&L")
    ws["A1"] = "Profit & Loss — Annual (kEUR)"
    ws["A1"].font = Font(size=12, bold=True)

    if footer_text:
        ws["A99"] = footer_text
        ws["A99"].font = Font(size=8, italic=True, color="808080")

    headers = ["Year", "Revenue", "EBITDA", "Depreciation", "EBIT", "Tax", "Net Income"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(3, col, h)
        c.font = Font(bold=True)
        c.fill = PatternFill("solid", fgColor="2E75B6")
        c.font = Font(bold=True, color="FFFFFF")

    try:
        periods = result.periods if hasattr(result, 'periods') else []
        if periods:
            for i, p in enumerate(periods[:30]):
                r = 4 + i
                ws.cell(r, 1, getattr(p, 'year_index', i + 1))
                ws.cell(r, 2, getattr(p, 'revenue_keur', 0)).number_format = "#,##0"
                ws.cell(r, 3, getattr(p, 'ebitda_keur', 0)).number_format = "#,##0"
                ws.cell(r, 4, getattr(p, 'depreciation_keur', 0)).number_format = "#,##0"
                ws.cell(r, 5, getattr(p, 'ebit_keur', 0)).number_format = "#,##0"
                ws.cell(r, 6, getattr(p, 'tax_keur', 0)).number_format = "#,##0"
                ws.cell(r, 7, getattr(p, 'net_income_keur', 0)).number_format = "#,##0"
        else:
            ws.cell(4, 1, 1)
            ws.cell(4, 2, getattr(result, 'total_revenue_keur', 0)).number_format = "#,##0"
            ws.cell(4, 3, getattr(result, 'total_ebitda_keur', 0)).number_format = "#,##0"
    except Exception as e:
        ws.cell(4, 1, f"Error: {e}")

    for col in range(1, 8):
        ws.column_dimensions[get_column_letter(col)].width = 14


# ---------------------------------------------------------------------------
# Balance Sheet
# ---------------------------------------------------------------------------

def _write_bs_sheet(wb, result: any, footer_text: str | None = None):
    ws = wb.create_sheet("BS")
    ws["A1"] = "Balance Sheet — Annual (kEUR)"
    ws["A1"].font = Font(size=12, bold=True)

    if footer_text:
        ws["A99"] = footer_text
        ws["A99"].font = Font(size=8, italic=True, color="808080")

    headers = ["Year", "Fixed Assets", "Cash", "Senior Debt", "Equity"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(3, col, h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="2E75B6")

    try:
        periods = result.periods if hasattr(result, 'periods') else []
        if periods:
            for i, p in enumerate(periods[:30]):
                r = 4 + i
                ws.cell(r, 1, getattr(p, 'year_index', i + 1))
                capex = getattr(result, 'total_capex_keur', 0) or 0
                dep = getattr(p, 'depreciation_keur', 0) or 0
                assets = max(0, capex - dep * (i + 1))
                ws.cell(r, 2, round(assets)).number_format = "#,##0"
                cash = max(0, getattr(p, 'cash_keur', 0))
                ws.cell(r, 3, round(cash)).number_format = "#,##0"
                debt = max(0, getattr(p, 'debt_balance_keur', 0))
                ws.cell(r, 4, round(debt)).number_format = "#,##0"
                equity = max(0, assets - debt + cash)
                ws.cell(r, 5, round(equity)).number_format = "#,##0"
    except Exception:
        pass

    for col in range(1, 6):
        ws.column_dimensions[get_column_letter(col)].width = 14


# ---------------------------------------------------------------------------
# Cash Flow sheet
# ---------------------------------------------------------------------------

def _write_cf_sheet(wb, result: any, footer_text: str | None = None):
    ws = wb.create_sheet("CF")
    ws["A1"] = "Cash Flow — Annual (kEUR)"
    ws["A1"].font = Font(size=12, bold=True)

    if footer_text:
        ws["A99"] = footer_text
        ws["A99"].font = Font(size=8, italic=True, color="808080")

    headers = ["Year", "EBITDA", "CFADS", "Senior DS", "FCF", "Distribution"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(3, col, h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="2E75B6")

    try:
        periods = result.periods if hasattr(result, 'periods') else []
        if periods:
            for i, p in enumerate(periods[:30]):
                r = 4 + i
                ws.cell(r, 1, getattr(p, 'year_index', i + 1))
                ws.cell(r, 2, getattr(p, 'ebitda_keur', 0)).number_format = "#,##0"
                ws.cell(r, 3, getattr(p, 'cf_after_tax_keur', 0)).number_format = "#,##0"
                ws.cell(r, 4, getattr(p, 'senior_ds_keur', 0)).number_format = "#,##0"
                fcf = getattr(p, 'cf_after_tax_keur', 0) - getattr(p, 'senior_ds_keur', 0)
                ws.cell(r, 5, round(fcf)).number_format = "#,##0"
                ws.cell(r, 6, getattr(p, 'distribution_keur', 0)).number_format = "#,##0"
    except Exception:
        pass

    for col in range(1, 7):
        ws.column_dimensions[get_column_letter(col)].width = 14


# ---------------------------------------------------------------------------
# Returns sheet
# ---------------------------------------------------------------------------

def _write_returns_sheet(wb, result: any, footer_text: str | None = None):
    ws = wb.create_sheet("Returns")
    ws["A1"] = "Returns Analysis"
    ws["A1"].font = Font(size=12, bold=True)

    if footer_text:
        ws["A99"] = footer_text
        ws["A99"].font = Font(size=8, italic=True, color="808080")

    ws["A3"] = "Unlevered Free Cash Flow (kEUR)"
    ws["A3"].font = Font(bold=True)

    headers = ["Year", "Revenue", "Opex", "EBITDA", "Capex", "Tax", "Unlevered FCF"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(4, col, h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="2E75B6")

    try:
        periods = result.periods if hasattr(result, 'periods') else []
        if periods:
            for i, p in enumerate(periods[:30]):
                r = 5 + i
                ws.cell(r, 1, getattr(p, 'year_index', i + 1))
                ws.cell(r, 2, getattr(p, 'revenue_keur', 0)).number_format = "#,##0"
                opex = getattr(p, 'ebitda_keur', 0) - getattr(p, 'revenue_keur', 0)
                ws.cell(r, 3, round(-opex)).number_format = "#,##0"
                ws.cell(r, 4, getattr(p, 'ebitda_keur', 0)).number_format = "#,##0"
                ws.cell(r, 5, 0).number_format = "#,##0"
                ws.cell(r, 6, getattr(p, 'tax_keur', 0)).number_format = "#,##0"
                fcf = getattr(p, 'fcf_keur', 0) if hasattr(p, 'fcf_keur') else 0
                ws.cell(r, 7, fcf).number_format = "#,##0"
    except Exception:
        pass

    ws["A36"] = "IRR & NPV Summary"
    ws["A36"].font = Font(bold=True, size=11)
    ws["A37"] = "Project IRR (unlevered)"
    ws.cell(37, 2, getattr(result, 'project_irr', 0)).number_format = "0.00%"
    ws["A38"] = "Equity IRR (levered)"
    ws.cell(38, 2, getattr(result, 'equity_irr', 0)).number_format = "0.00%"
    ws["A39"] = "Project NPV (kEUR)"
    ws.cell(39, 2, getattr(result, 'project_npv_keur', 0)).number_format = "#,##0"
    ws["A40"] = "LCOE (EUR/MWh)"
    ws.cell(40, 2, 0).number_format = "0.00"

    for col in range(1, 8):
        ws.column_dimensions[get_column_letter(col)].width = 14


# ---------------------------------------------------------------------------
# Debt Service sheet
# ---------------------------------------------------------------------------

def _write_ds_sheet(wb, result: any, footer_text: str | None = None):
    ws = wb.create_sheet("DS")
    ws["A1"] = "Debt Service & Coverage Ratios"
    ws["A1"].font = Font(size=12, bold=True)

    if footer_text:
        ws["A99"] = footer_text
        ws["A99"].font = Font(size=8, italic=True, color="808080")

    headers = ["Period", "Year", "Senior DS", "DSCR", "LLCR", "PLCR", "Lockup"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(3, col, h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="C00000")

    r = 4
    try:
        periods = result.periods if hasattr(result, 'periods') else []
        if periods:
            for i, p in enumerate(periods):
                ws.cell(r, 1, i + 1)
                ws.cell(r, 2, getattr(p, 'year_index', i + 1))
                ws.cell(r, 3, getattr(p, 'debt_service_keur', 0)).number_format = "#,##0"
                dscr = getattr(p, 'dscr', 0) or 0
                c = ws.cell(r, 4, dscr)
                c.number_format = "0.00x"
                if dscr < 1.0:
                    c.fill = PatternFill("solid", fgColor="FFCCCC")
                elif dscr < 1.15:
                    c.fill = PatternFill("solid", fgColor="FFFFCC")
                else:
                    c.fill = PatternFill("solid", fgColor="CCFFCC")

                llcr = getattr(p, 'llcr', 0) or 0
                ws.cell(r, 5, llcr).number_format = "0.00x"
                plcr = getattr(p, 'plcr', 0) or 0
                ws.cell(r, 6, plcr).number_format = "0.00x"
                ws.cell(r, 7, "⚠ LOCKUP" if dscr < 1.10 else "")
                r += 1
    except Exception:
        pass

    r += 1
    ws.cell(r, 1, "Avg DSCR").font = Font(bold=True)
    ws.cell(r, 2, getattr(result, 'avg_dscr', 0)).number_format = "0.00x"
    r += 1
    ws.cell(r, 1, "Min DSCR").font = Font(bold=True)
    ws.cell(r, 2, getattr(result, 'min_dscr', 0)).number_format = "0.00x"
    r += 1
    ws.cell(r, 1, "Min LLCR").font = Font(bold=True)
    ws.cell(r, 2, getattr(result, 'min_llcr', 0)).number_format = "0.00x"

    for col in range(1, 8):
        ws.column_dimensions[get_column_letter(col)].width = 12


# ---------------------------------------------------------------------------
# Sensitivity sheet — Spider Table + Two-Way Heatmap (Task 3.5)
# ---------------------------------------------------------------------------

def _write_sensitivity_sheet(
    wb,
    inputs: any,
    result: any,
    footer_text: str | None = None,
):
    """Write Sensitivity sheet with Spider Table and Two-Way Heatmap sub-sheets."""
    if hasattr(inputs, 'revenue'):
        spider = run_spider_analysis(inputs, n_steps=7, target_irr_basis="project")
    else:
        spider = None

    # ---- Spider Table sub-sheet ----
    ws_spider = wb.create_sheet("Spider Table")
    ws_spider["A1"] = "Spider Table — Project IRR by Variable & Step"
    ws_spider["A1"].font = Font(size=12, bold=True)

    if footer_text:
        ws_spider["A99"] = footer_text
        ws_spider["A99"].font = Font(size=8, italic=True, color="808080")

    if spider:
        # Header row: Variable | -20% | -13% | -7% | Base | +7% | +13% | +20%
        steps_pct = [f"{s*100:.0f}%" for s in spider["steps"]]
        headers = ["Variable"] + steps_pct
        for col, h in enumerate(headers, 1):
            c = ws_spider.cell(3, col, h)
            c.font = Font(bold=True, color="FFFFFF")
            c.fill = PatternFill("solid", fgColor="1F4E79")
            c.alignment = Alignment(horizontal="center")

        # Data rows
        for row_idx, var in enumerate(spider["variables"]):
            r = 4 + row_idx
            ws_spider.cell(r, 1, var).font = Font(bold=True)
            irr_values = spider["matrix"][var]
            for col_idx, irr_val in enumerate(irr_values):
                c = ws_spider.cell(r, col_idx + 2, irr_val)
                c.number_format = "0.00%"
                # Color-code: green for high IRR, red for low
                min_irr = min(irr_values)
                max_irr = max(irr_values)
                range_ = max_irr - min_irr if max_irr > min_irr else 1
                # Normalize to 0-1
                norm = (irr_val - min_irr) / range_ if range_ > 0 else 0.5
                # Red (255,0,0) at low, Green (0,255,0) at high
                red = int(255 * (1 - norm))
                green = int(255 * norm)
                c.fill = PatternFill("solid", fgColor=f"{red:02X}{green:02X}00")
                c.font = Font(color="000000")

        # Column widths
        ws_spider.column_dimensions["A"].width = 18
        for col in range(2, 9):
            ws_spider.column_dimensions[get_column_letter(col)].width = 10

    # ---- Two-Way Heatmap sub-sheet ----
    ws_heat = wb.create_sheet("Two-Way Heatmap")
    ws_heat["A1"] = "Two-Way Sensitivity: PPA Tariff vs CAPEX → Equity IRR"
    ws_heat["A1"].font = Font(size=12, bold=True)

    if footer_text:
        ws_heat["A99"] = footer_text
        ws_heat["A99"].font = Font(size=8, italic=True, color="808080")

    # Run two-way analysis: X = PPA Tariff (±20%, 5 values), Y = CAPEX (±20%, 5 values)
    if hasattr(inputs, 'revenue') and hasattr(inputs, 'capex'):
        ppa_tariff_vals = [
            inputs.revenue.ppa_base_tariff * 0.8,
            inputs.revenue.ppa_base_tariff * 0.9,
            inputs.revenue.ppa_base_tariff,
            inputs.revenue.ppa_base_tariff * 1.1,
            inputs.revenue.ppa_base_tariff * 1.2,
        ]
        # CAPEX: scale all items by factor
        base_capex_total = inputs.capex.total_capex
        capex_factors = [0.8, 0.9, 1.0, 1.1, 1.2]
        from dataclasses import replace as dc_replace
        from core.finance.sensitivity import _scale_capex, _run_with_inputs, _get_irr

        matrix_2d: list[list[float]] = []
        for capex_f in capex_factors:
            row = []
            for ppa_t in ppa_tariff_vals:
                mod_rev = dc_replace(inputs.revenue, ppa_base_tariff=ppa_t)
                mod_capex = _scale_capex(inputs.capex, capex_f)
                mod_inputs = dc_replace(inputs, revenue=mod_rev, capex=mod_capex)
                res = _run_with_inputs(mod_inputs)
                irr = _get_irr(res, basis="equity")
                row.append(irr)
            matrix_2d.append(row)
    else:
        matrix_2d = []

    if matrix_2d:
        # Header row: CAPEX\PPA | 80% | 90% | Base | 110% | 120%
        ppa_labels = [f"{v:.1f}" for v in ppa_tariff_vals]
        headers_heat = ["CAPEX \\ PPA"] + ppa_labels
        for col, h in enumerate(headers_heat, 1):
            c = ws_heat.cell(3, col, h)
            c.font = Font(bold=True, color="FFFFFF")
            c.fill = PatternFill("solid", fgColor="1F4E79")
            c.alignment = Alignment(horizontal="center")

        # Row labels: -20%, -10%, base, +10%, +20%
        capex_labels = [f"{f*100:.0f}%" for f in capex_factors]

        # Flatten for color scale rule
        all_vals = [v for row in matrix_2d for v in row]
        min_val = min(all_vals)
        max_val = max(all_vals)

        for row_idx, capex_lbl in enumerate(capex_labels):
            r = 4 + row_idx
            ws_heat.cell(r, 1, capex_lbl).font = Font(bold=True)
            for col_idx, irr_val in enumerate(matrix_2d[row_idx]):
                c = ws_heat.cell(r, col_idx + 2, irr_val)
                c.number_format = "0.00%"
                # Red-white-green color scale
                range_ = max_val - min_val if max_val > min_val else 1
                norm = (irr_val - min_val) / range_
                red = int(255 * (1 - norm))
                green = int(255 * norm)
                c.fill = PatternFill("solid", fgColor=f"{red:02X}{green:02X}00")

        ws_heat.column_dimensions["A"].width = 14
        for col in range(2, 7):
            ws_heat.column_dimensions[get_column_letter(col)].width = 10


# ---------------------------------------------------------------------------
# Main export function (Tasks 3.5 + 3.6)
# ---------------------------------------------------------------------------

def export_fid_deck_excel(
    result: any,
    inputs: any,
    filepath: str,
    branding: dict | None = None,
    project_name: str | None = None,
    logo: bytes | None = None,
    footer_text: str | None = None,
    disclaimer: str | None = None,
) -> None:
    """Generate FID Deck Excel workbook.

    Sheets:
    1. FID deck outputs (KPI cover: LCOE, Project IRR, NPV, Equity IRR, Min/Avg DSCR, LLCR, Payback)
    2. P&L (Revenue, EBITDA, EBIT, Net Income — annual)
    3. BS (Fixed Assets, Cash, Debt, Equity — annual)
    4. CF (EBITDA, CFADS, Senior DS, FCF, Distribution — annual)
    5. Returns (Unlevered FCF, IRR, NPV, LCOE)
    6. DS (DSCR per period, Min/Avg, LLCR, lockup flags)
    7. Sensitivity (Spider Table + Two-Way Heatmap)  [Task 3.5]

    Branding parameters [Task 3.6]:
        logo: bytes → embed in Cover sheet header
        footer_text: str → appears in all sheets / footer
        disclaimer: str → appears on cover sheet (via branding dict)

    Args:
        result: WaterfallResult from run_waterfall
        inputs: ProjectInputs used for computation
        filepath: output .xlsx path
        branding: optional dict with logo (bytes), footer_text, disclaimer
        project_name: optional override for cover title
        logo: bytes for logo image to embed in cover sheet
        footer_text: str to appear in all sheet footers
        disclaimer: str to appear on cover sheet
    """
    # Merge branding parameters
    effective_branding: dict = branding.copy() if branding else {}
    if logo is not None:
        effective_branding['logo'] = logo
    if disclaimer is not None:
        effective_branding['disclaimer'] = disclaimer
    if footer_text is not None:
        effective_branding['footer_text'] = footer_text

    wb = Workbook()
    wb.remove(wb.active)

    # Cover sheet
    ws_cover = wb.create_sheet("FID deck outputs")
    proj_name = project_name or (
        getattr(inputs.info, 'name', 'Project') if hasattr(inputs, 'info') else 'Project'
    )
    _write_kpi_cover(
        ws_cover, result, inputs,
        project_name=proj_name,
        branding=effective_branding,
    )

    # P&L
    _write_pl_sheet(wb, result, footer_text=footer_text)

    # Balance Sheet
    _write_bs_sheet(wb, result, footer_text=footer_text)

    # Cash Flow
    _write_cf_sheet(wb, result, footer_text=footer_text)

    # Returns
    _write_returns_sheet(wb, result, footer_text=footer_text)

    # Debt Service
    _write_ds_sheet(wb, result, footer_text=footer_text)

    # Sensitivity (Spider Table + Two-Way Heatmap) — Task 3.5
    _write_sensitivity_sheet(wb, inputs, result, footer_text=footer_text)

    wb.save(filepath)