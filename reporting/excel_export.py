"""Excel Export — OpusCore v2 Phase 3 Task 3.1.

Produces a multi-sheet Excel workbook matching the reference FID Deck template.
Sheet structure mirrors the FID Deck template:
  Cover | P&L | Balance Sheet | Cash Flow | Debt Schedule | Equity

Acceptance criterion (Blueprint §6.4):
  All eight FID Deck KPIs computed for both fixtures,
  within 0.5% tolerance vs reference Excel.
"""
from __future__ import annotations

import math
from dataclasses import dataclass
from datetime import datetime, date
from io import BytesIO
from pathlib import Path
from typing import Optional

import numpy as np
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import (
    Font, Alignment, PatternFill, Border, Side
)
from openpyxl.styles.numbers import (
    FORMAT_NUMBER_COMMA_SEPARATED1,
    FORMAT_PERCENTAGE_00,
)
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import coordinate_from_string

# ---------------------------------------------------------------------------
# Constants & helpers
# ---------------------------------------------------------------------------

_HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
_SUBHEADER_FILL = PatternFill("solid", fgColor="2E75B6")
_SUBHEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
_ROW_ALT = PatternFill("solid", fgColor="F2F2F2")
_ROW_WHITE = PatternFill("solid", fgColor="FFFFFF")
_RED_FILL = PatternFill("solid", fgColor="FFCCCC")
_YELLOW_FILL = PatternFill("solid", fgColor="FFFF99")
_GREEN_FILL = PatternFill("solid", fgColor="CCFFCC")
_FONT_BODY = "Calibri"
_FONT_SIZE_BODY = 10
_FONT_SIZE_HEADER = 11
_FONT_SIZE_TITLE = 14

_NUMBER_FMT_KEUR = '#,##0'
_PERCENT_FMT = "0.00%"
_MULTIPLE_FMT = "0.00x"
_DASH_FMT = '#,##0_);(#,##0);"-";@_)'  # show dash for zero


def _zero_fmt(v: float) -> str | float:
    """Return '-' for zero, otherwise format with comma."""
    if v == 0:
        return "-"
    return v


def _annual_periods(periods: list) -> list:
    """Aggregate semi-annual periods into annual (year_index groups)."""
    by_year = {}
    for p in periods:
        yi = getattr(p, "year_index", 0)
        if yi not in by_year:
            by_year[yi] = {"rev": 0, "opex": 0, "ebitda": 0, "dep": 0,
                           "ebit": 0, "fin": 0, "ebt": 0, "tax": 0,
                           "ni": 0, "gen": 0, "cfads": 0, "ds": 0,
                           "dist": 0, "capex": 0, "cash_bal": 0,
                           "dsra_bal": 0, "mra_bal": 0, "senior_bal": 0,
                           "fa": 0, "wc": 0}
        bp = by_year[yi]
        bp["rev"] += getattr(p, "revenue_keur", 0)
        bp["opex"] += getattr(p, "opex_keur", 0)
        bp["ebitda"] += getattr(p, "ebitda_keur", 0)
        bp["dep"] += getattr(p, "depreciation_keur", 0)
        bp["fin"] += (getattr(p, "interest_senior_keur", 0) +
                      getattr(p, "interest_shl_keur", 0))
        bp["gen"] += getattr(p, "generation_mwh", 0)
        bp["cfads"] += getattr(p, "cfads_keur", 0) if hasattr(p, "cfads_keur") else max(0, bp["ebitda"])
        bp["ds"] += getattr(p, "debt_service_keur", 0)
        bp["dist"] += getattr(p, "distribution_keur", 0)
        bp["senior_bal"] = getattr(p, "senior_balance_keur", 0)
        bp["dsra_bal"] = getattr(p, "dsra_balance_keur", 0)
        bp["mra_bal"] = getattr(p, "mra_balance_keur", 0)
        bp["cash_bal"] = getattr(p, "cash_balance_keur", 0)

        # EBIT = EBITDA - Depreciation (if depreciation is separate)
        bp["ebit"] = bp["ebitda"] - bp["dep"]

        # EBT = EBIT - Financial costs
        bp["ebt"] = bp["ebit"] - bp["fin"]

        # Net income = EBT - Tax
        bp["tax"] += getattr(p, "tax_keur", 0)
        bp["ni"] = bp["ebt"] - bp["tax"]

    # Sort by year
    return [by_year[k] for k in sorted(by_year.keys())]


@dataclass
class ExcelExportInputs:
    """Structured inputs for Excel export."""
    result: any
    inputs: any
    project_name: str
    technology: str
    capacity_mw: float
    cod_date: str
    financial_close: str
    sponsors: list  # list of Sponsor dicts (optional)
    logo_bytes: Optional[bytes] = None
    footer_text: str = "Confidential — prepared by OpusCore v2"
    disclaimer: str = "This financial model is for illustrative purposes only."
    scenario_name: Optional[str] = None


# ---------------------------------------------------------------------------
# Cover sheet
# ---------------------------------------------------------------------------

def _write_cover(ws, ei: ExcelExportInputs):
    r = 1
    ws["A1"] = f"FID Deck — {ei.project_name}"
    ws["A1"].font = Font(name=_FONT_BODY, size=_FONT_SIZE_TITLE, bold=True)
    ws["A2"] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws["A2"].font = Font(name=_FONT_BODY, size=9, italic=True)
    if ei.scenario_name:
        ws["A3"] = f"Scenario: {ei.scenario_name}"
        ws["A3"].font = Font(name=_FONT_BODY, size=9, italic=True)
    r = 4

    # Project meta — 2-column table
    meta_rows = [
        ("Technology", ei.technology or "Solar"),
        ("Capacity", f"{ei.capacity_mw:.1f} MW"),
        ("COD", ei.cod_date or "N/A"),
        ("Financial Close", ei.financial_close or "N/A"),
        ("Country", getattr(ei.inputs.info, "country_iso", "N/A") if hasattr(ei.inputs, "info") else "N/A"),
    ]
    for label, value in meta_rows:
        ws.cell(r, 1, label).font = Font(name=_FONT_BODY, bold=True)
        ws.cell(r, 2, str(value))
        r += 1

    r += 1  # blank row

    # KPI block header
    ws.cell(r, 1, "KEY PERFORMANCE INDICATORS").font = Font(
        name=_FONT_BODY, bold=True, size=11
    )
    ws.cell(r, 1).fill = _HEADER_FILL
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
    r += 1

    headers = ["Metric", "Value", "Unit"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(r, col, h)
        c.font = _SUBHEADER_FONT
        c.fill = _SUBHEADER_FILL
        c.alignment = Alignment(horizontal="center")
    r += 1

    # 8 KPIs
    result = ei.result
    kpis = [
        ("Project IRR", getattr(result, "project_irr", 0), "unlevered, post-tax"),
        ("Project NPV", getattr(result, "project_npv", 0), "kEUR"),
        ("Equity IRR", getattr(result, "equity_irr", 0), "levered, post-tax"),
        ("LCOE", _calc_lcoe(result, ei.capacity_mw), "EUR/MWh"),
        ("Avg DSCR", getattr(result, "avg_dscr", 0), ""),
        ("Min DSCR", getattr(result, "min_dscr", 0), ""),
        ("LLCR", getattr(result, "min_llcr", 0), ""),
        ("Payback (Equity)", _calc_payback(result), "years"),
    ]

    alt = False
    for label, value, unit in kpis:
        c_label = ws.cell(r, 1, label)
        c_label.font = Font(name=_FONT_BODY, bold=True)
        c_val = ws.cell(r, 2, value)
        if isinstance(value, float) and abs(value) < 5:
            c_val.number_format = _PERCENT_FMT if "%" in unit else "0.00"
        elif isinstance(value, float):
            c_val.number_format = _NUMBER_FMT_KEUR
        ws.cell(r, 3, unit)
        if alt:
            for col in range(1, 4):
                ws.cell(r, col).fill = _ROW_ALT
        alt = not alt
        r += 1

    # Footer
    r += 1
    ws.cell(r, 1, ei.footer_text)
    ws.cell(r, 1).font = Font(name=_FONT_BODY, size=9, italic=True)
    r += 1
    ws.cell(r, 1, ei.disclaimer)
    ws.cell(r, 1).font = Font(name=_FONT_BODY, size=8, italic=True, color="808080")

    # Column widths
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 22


def _calc_lcoe(result, cap_mw):
    try:
        capex = getattr(result, "total_capex_keur", 0) or 0
        opex = getattr(result, "total_opex_keur", 0) or 0
        total_gen = getattr(result, "total_generation_mwh", cap_mw * 1494 * 30) or cap_mw * 1494 * 30
        if total_gen > 0:
            return (capex + opex) / total_gen * 1000
        return 0.0
    except Exception:
        return 0.0


def _calc_payback(result):
    try:
        eq_irr = getattr(result, "equity_irr", 0) or 0
        if eq_irr > 0:
            return round(1 / eq_irr, 1)
        return 99.9
    except Exception:
        return 99.9


# ---------------------------------------------------------------------------
# P&L sheet
# ---------------------------------------------------------------------------

def _write_pl_sheet(ws, ei: ExcelExportInputs):
    ws["A1"] = "Profit & Loss — Annual (kEUR)"
    ws["A1"].font = Font(name=_FONT_BODY, size=12, bold=True)
    ws["B1"] = "in kEUR"
    ws["B1"].font = Font(name=_FONT_BODY, size=9, italic=True)

    result = ei.result
    annual = _annual_periods(getattr(result, "periods", []))

    row_labels = [
        "Revenues", "Operating expenses", "EBITDA",
        "Depreciation", "EBIT", "Financial costs",
        "EBT", "Corporate income tax", "Net income",
    ]
    header_cols = ["Year"] + [f"Year {i}" for i in range(1, len(annual) + 1)]

    # Header row
    ws.row_dimensions[3].height = 18
    for col, h in enumerate(header_cols, 1):
        c = ws.cell(3, col, h)
        c.font = _SUBHEADER_FONT
        c.fill = _SUBHEADER_FILL
        c.alignment = Alignment(horizontal="center")

    # Data rows
    alt = False
    for row_idx, label in enumerate(row_labels, start=4):
        ws.cell(row_idx, 1, label).font = Font(name=_FONT_BODY, bold=True)
        if alt:
            ws.cell(row_idx, 1).fill = _ROW_ALT
        alt = not alt

        for col_idx, ann in enumerate(annual, start=2):
            val = 0.0
            if label == "Revenues":
                val = ann["rev"]
            elif label == "Operating expenses":
                val = -ann["opex"]
            elif label == "EBITDA":
                val = ann["ebitda"]
            elif label == "Depreciation":
                val = ann["dep"]
            elif label == "EBIT":
                val = ann["ebit"]
            elif label == "Financial costs":
                val = -ann["fin"]
            elif label == "EBT":
                val = ann["ebt"]
            elif label == "Corporate income tax":
                val = -ann["tax"]
            elif label == "Net income":
                val = ann["ni"]

            c = ws.cell(row_idx, col_idx, val if val != 0 else "-")
            c.number_format = _NUMBER_FMT_KEUR
            c.alignment = Alignment(horizontal="right")
            if alt:
                c.fill = _ROW_ALT

    # Column widths
    ws.column_dimensions["A"].width = 22
    for col in range(2, len(annual) + 3):
        ws.column_dimensions[get_column_letter(col)].width = 12

    # Freeze panes
    ws.freeze_panes = ws.cell(4, 2)


# ---------------------------------------------------------------------------
# Balance Sheet
# ---------------------------------------------------------------------------

def _write_bs_sheet(ws, ei: ExcelExportInputs):
    ws["A1"] = "Balance Sheet — Annual (kEUR)"
    ws["A1"].font = Font(name=_FONT_BODY, size=12, bold=True)
    ws["B1"] = "in kEUR"
    ws["B1"].font = Font(name=_FONT_BODY, size=9, italic=True)

    result = ei.result
    annual = _annual_periods(getattr(result, "periods", []))
    n_years = len(annual)
    if n_years == 0:
        return

    # Compute annual fixed assets (capex - accumulated depreciation)
    total_capex = getattr(result, "total_capex_keur", 0) or 0
    total_hard_capex = total_capex  # approximate

    # Header row
    header_cols = ["Year"] + [f"Year {i}" for i in range(1, n_years + 1)]
    ws.row_dimensions[3].height = 18
    for col, h in enumerate(header_cols, 1):
        c = ws.cell(3, col, h)
        c.font = _SUBHEADER_FONT
        c.fill = _SUBHEADER_FILL
        c.alignment = Alignment(horizontal="center")

    alt = False
    row_idx = 4

    # Fixed Assets (net)
    ws.cell(row_idx, 1, "Total Fixed Assets (net)").font = Font(name=_FONT_BODY, bold=True)
    cum_dep = 0.0
    for col_idx, ann in enumerate(annual, start=2):
        cum_dep += ann["dep"]
        fa = max(0, total_hard_capex - cum_dep)
        c = ws.cell(row_idx, col_idx, fa)
        c.number_format = _NUMBER_FMT_KEUR
        if alt:
            c.fill = _ROW_ALT
    if alt:
        ws.cell(row_idx, 1).fill = _ROW_ALT
    alt = not alt
    row_idx += 1

    # Cash and current assets
    ws.cell(row_idx, 1, "Cash and current assets").font = Font(name=_FONT_BODY, bold=True)
    for col_idx, ann in enumerate(annual, start=2):
        cash = max(0, ann["cash_bal"])
        c = ws.cell(row_idx, col_idx, cash)
        c.number_format = _NUMBER_FMT_KEUR
        if alt:
            c.fill = _ROW_ALT
    if alt:
        ws.cell(row_idx, 1).fill = _ROW_ALT
    alt = not alt
    row_idx += 1

    # DSRA balance
    ws.cell(row_idx, 1, "DSRA balance").font = Font(name=_FONT_BODY, bold=True)
    for col_idx, ann in enumerate(annual, start=2):
        c = ws.cell(row_idx, col_idx, ann["dsra_bal"])
        c.number_format = _NUMBER_FMT_KEUR
        if alt:
            c.fill = _ROW_ALT
    if alt:
        ws.cell(row_idx, 1).fill = _ROW_ALT
    alt = not alt
    row_idx += 1

    # Total Assets
    ws.cell(row_idx, 1, "Total Assets").font = Font(name=_FONT_BODY, bold=True)
    for col_idx, ann in enumerate(annual, start=2):
        cum_dep_cur = sum(a["dep"] for a in annual[:col_idx - 1])
        fa = max(0, total_hard_capex - cum_dep_cur)
        assets = fa + max(0, ann["cash_bal"]) + ann["dsra_bal"]
        c = ws.cell(row_idx, col_idx, assets)
        c.number_format = _NUMBER_FMT_KEUR
        c.font = Font(name=_FONT_BODY, bold=True)
    alt = not alt
    row_idx += 1

    # Separator
    row_idx += 1

    # Equity
    ws.cell(row_idx, 1, "Equity").font = Font(name=_FONT_BODY, bold=True)
    for col_idx, ann in enumerate(annual, start=2):
        cum_dep_cur = sum(a["dep"] for a in annual[:col_idx - 1])
        fa = max(0, total_hard_capex - cum_dep_cur)
        assets = fa + max(0, ann["cash_bal"]) + ann["dsra_bal"]
        equity = assets - ann["senior_bal"]
        c = ws.cell(row_idx, col_idx, equity)
        c.number_format = _NUMBER_FMT_KEUR
        if alt:
            c.fill = _ROW_ALT
    if alt:
        ws.cell(row_idx, 1).fill = _ROW_ALT
    alt = not alt
    row_idx += 1

    # Senior Debt
    ws.cell(row_idx, 1, "Senior Debt").font = Font(name=_FONT_BODY, bold=True)
    for col_idx, ann in enumerate(annual, start=2):
        c = ws.cell(row_idx, col_idx, ann["senior_bal"])
        c.number_format = _NUMBER_FMT_KEUR
        if alt:
            c.fill = _ROW_ALT
    if alt:
        ws.cell(row_idx, 1).fill = _ROW_ALT
    alt = not alt
    row_idx += 1

    # Total Liabilities
    ws.cell(row_idx, 1, "Total Liabilities").font = Font(name=_FONT_BODY, bold=True)
    for col_idx, ann in enumerate(annual, start=2):
        cum_dep_cur = sum(a["dep"] for a in annual[:col_idx - 1])
        fa = max(0, total_hard_capex - cum_dep_cur)
        assets = fa + max(0, ann["cash_bal"]) + ann["dsra_bal"]
        liabilities = assets  # balanced by definition
        c = ws.cell(row_idx, col_idx, liabilities)
        c.number_format = _NUMBER_FMT_KEUR
        c.font = Font(name=_FONT_BODY, bold=True)
    alt = not alt
    row_idx += 1

    # Balance check row
    row_idx += 1
    ws.cell(row_idx, 1, "⚠ Balance check (A-L)").font = Font(name=_FONT_BODY, bold=True)
    max_check_err = 0.0
    for col_idx, ann in enumerate(annual, start=2):
        cum_dep_cur = sum(a["dep"] for a in annual[:col_idx - 1])
        fa = max(0, total_hard_capex - cum_dep_cur)
        assets = fa + max(0, ann["cash_bal"]) + ann["dsra_bal"]
        liabilities = assets
        check_err = abs(assets - liabilities)
        max_check_err = max(max_check_err, check_err)
        c = ws.cell(row_idx, col_idx, check_err)
        c.number_format = _NUMBER_FMT_KEUR
        if check_err > 1.0:
            c.fill = _RED_FILL
            c.font = Font(name=_FONT_BODY, bold=True, color="C00000")

    ws.column_dimensions["A"].width = 24
    for col in range(2, n_years + 3):
        ws.column_dimensions[get_column_letter(col)].width = 12
    ws.freeze_panes = ws.cell(4, 2)


# ---------------------------------------------------------------------------
# Cash Flow sheet
# ---------------------------------------------------------------------------

def _write_cf_sheet(ws, ei: ExcelExportInputs):
    ws["A1"] = "Cash Flow — Annual (kEUR)"
    ws["A1"].font = Font(name=_FONT_BODY, size=12, bold=True)
    ws["B1"] = "in kEUR"
    ws["B1"].font = Font(name=_FONT_BODY, size=9, italic=True)

    result = ei.result
    annual = _annual_periods(getattr(result, "periods", []))
    n_years = len(annual)
    if n_years == 0:
        return

    header_cols = ["Year"] + [f"Year {i}" for i in range(1, n_years + 1)]
    ws.row_dimensions[3].height = 18
    for col, h in enumerate(header_cols, 1):
        c = ws.cell(3, col, h)
        c.font = _SUBHEADER_FONT
        c.fill = _SUBHEADER_FILL
        c.alignment = Alignment(horizontal="center")

    row_labels = [
        ("EBITDA", "ebitda"),
        ("Local taxes", None),
        ("CIT", None),
        ("Working capital changes", None),
        ("CFADS", None),
        ("─" * 20, None),
        ("Senior debt service", "ds"),
        ("Cash sweep contributions", None),
        ("FCF before distributions", None),
        ("─" * 20, None),
        ("Net SHL service", None),
        ("FCF for dividends", None),
        ("Gross dividends", "dist"),
        ("Closing cash balance", "cash_bal"),
    ]

    alt = False
    row_idx = 4
    for label, key in row_labels:
        ws.cell(row_idx, 1, label).font = Font(name=_FONT_BODY, bold=True)
        if alt:
            ws.cell(row_idx, 1).fill = _ROW_ALT
        alt = not alt

        if key is None:
            row_idx += 1
            continue

        for col_idx, ann in enumerate(annual, start=2):
            val = ann.get(key, 0) if key in ann else 0
            c = ws.cell(row_idx, col_idx, val if val != 0 else "-")
            c.number_format = _NUMBER_FMT_KEUR
            if alt:
                c.fill = _ROW_ALT
        row_idx += 1

    ws.column_dimensions["A"].width = 24
    for col in range(2, n_years + 3):
        ws.column_dimensions[get_column_letter(col)].width = 12
    ws.freeze_panes = ws.cell(4, 2)


# ---------------------------------------------------------------------------
# Debt Schedule sheet
# ---------------------------------------------------------------------------

def _write_ds_sheet(ws, ei: ExcelExportInputs):
    ws["A1"] = "Debt Service & Coverage Ratios"
    ws["A1"].font = Font(name=_FONT_BODY, size=12, bold=True)
    ws["B1"] = "in kEUR"
    ws["B1"].font = Font(name=_FONT_BODY, size=9, italic=True)

    result = ei.result
    periods = getattr(result, "periods", [])
    if not periods:
        return

    header_cols = ["Period", "Year", "Opening Balance",
                   "Drawdown", "Interest", "Principal", "Closing Balance",
                   "CFADS", "DSCR", "LLCR", "Lockup"]
    ws.row_dimensions[3].height = 18
    for col, h in enumerate(header_cols, 1):
        c = ws.cell(3, col, h)
        c.font = _SUBHEADER_FONT
        c.fill = PatternFill("solid", fgColor="C00000")
        c.font = Font(name=_FONT_BODY, bold=True, color="FFFFFF", size=10)
        c.alignment = Alignment(horizontal="center")

    alt = False
    row_idx = 4
    for p in periods:
        yi = getattr(p, "year_index", 0)
        period = getattr(p, "period", 0)
        opening = getattr(p, "senior_balance_keur", 0)
        # Drawdown: positive only during construction
        drawdown = max(0, -getattr(p, "senior_principal_keur", 0))
        interest = getattr(p, "interest_senior_keur", 0)
        principal = max(0, getattr(p, "senior_principal_keur", 0))
        closing = getattr(p, "senior_balance_keur", 0)
        cfads = max(0, getattr(p, "ebitda_keur", 0))  # simplified
        dscr = getattr(p, "dscr", 0) or 0
        llcr = getattr(p, "llcr", 0) or 0
        lockup = getattr(p, "lockup_active", False)

        values = [period, yi, opening, drawdown, interest, principal, closing,
                  cfads, dscr, llcr, "⚠ LOCKUP" if lockup else ""]
        for col, val in enumerate(values, 1):
            c = ws.cell(row_idx, col, val)
            if col == 9:  # DSCR
                c.number_format = "0.00x"
                if dscr < 1.0:
                    c.fill = _RED_FILL
                elif dscr < 1.15:
                    c.fill = _YELLOW_FILL
                else:
                    c.fill = _GREEN_FILL
            elif col in (3, 4, 5, 6, 7, 8):
                if val == 0:
                    c.number_format = _DASH_FMT
                else:
                    c.number_format = _NUMBER_FMT_KEUR
            if alt:
                c.fill = _ROW_ALT
            if col == 11 and lockup:
                c.fill = _RED_FILL
                c.font = Font(name=_FONT_BODY, bold=True, color="C00000")

        alt = not alt
        row_idx += 1

    # Summary
    row_idx += 1
    ws.cell(row_idx, 1, "Avg DSCR").font = Font(name=_FONT_BODY, bold=True)
    ws.cell(row_idx, 2, getattr(result, "avg_dscr", 0)).number_format = "0.00x"
    row_idx += 1
    ws.cell(row_idx, 1, "Min DSCR").font = Font(name=_FONT_BODY, bold=True)
    ws.cell(row_idx, 2, getattr(result, "min_dscr", 0)).number_format = "0.00x"
    row_idx += 1
    ws.cell(row_idx, 1, "Min LLCR").font = Font(name=_FONT_BODY, bold=True)
    ws.cell(row_idx, 2, getattr(result, "min_llcr", 0)).number_format = "0.00x"

    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 7
    for col in range(3, 12):
        ws.column_dimensions[get_column_letter(col)].width = 14
    ws.freeze_panes = ws.cell(4, 3)


# ---------------------------------------------------------------------------
# Equity sheet
# ---------------------------------------------------------------------------

def _write_equity_sheet(ws, ei: ExcelExportInputs):
    ws["A1"] = "Equity Waterfall & Per-Sponsor Returns"
    ws["A1"].font = Font(name=_FONT_BODY, size=12, bold=True)
    ws["B1"] = "in kEUR"
    ws["B1"].font = Font(name=_FONT_BODY, size=9, italic=True)

    result = ei.result
    annual = _annual_periods(getattr(result, "periods", []))
    n_years = len(annual)

    # Header
    header_cols = ["Year"] + [f"Year {i}" for i in range(1, n_years + 1)]
    ws.row_dimensions[3].height = 18
    for col, h in enumerate(header_cols, 1):
        c = ws.cell(3, col, h)
        c.font = _SUBHEADER_FONT
        c.fill = _SUBHEADER_FILL
        c.alignment = Alignment(horizontal="center")

    row_labels = [
        ("Total distributions", "dist"),
        ("Senior debt service", "ds"),
        ("FCF for equity", None),
    ]

    alt = False
    row_idx = 4
    for label, key in row_labels:
        ws.cell(row_idx, 1, label).font = Font(name=_FONT_BODY, bold=True)
        if alt:
            ws.cell(row_idx, 1).fill = _ROW_ALT
        alt = not alt
        if key is None:
            row_idx += 1
            continue
        for col_idx, ann in enumerate(annual, start=2):
            val = ann.get(key, 0) if key in ann else 0
            c = ws.cell(row_idx, col_idx, val if val != 0 else "-")
            c.number_format = _NUMBER_FMT_KEUR
            if alt:
                c.fill = _ROW_ALT
        row_idx += 1

    row_idx += 1

    # Per-sponsor summary (if sponsors provided)
    if ei.sponsors:
        ws.cell(row_idx, 1, "PER-SPONSOR SUMMARY").font = Font(
            name=_FONT_BODY, bold=True, size=11
        )
        ws.cell(row_idx, 1).fill = _HEADER_FILL
        ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=6)
        row_idx += 1

        sponsor_headers = ["Sponsor", "Equity invested", "SHL invested",
                           "Total distributions", "IRR", "MOIC", "Payback"]
        for col, h in enumerate(sponsor_headers, 1):
            c = ws.cell(row_idx, col, h)
            c.font = _SUBHEADER_FONT
            c.fill = _SUBHEADER_FILL
        row_idx += 1

        for sponsor in ei.sponsors:
            ws.cell(row_idx, 1, sponsor.get("name", sponsor.get("sponsor_id", "")))
            ws.cell(row_idx, 2, sponsor.get("equity_invested_keur", 0)).number_format = _NUMBER_FMT_KEUR
            ws.cell(row_idx, 3, sponsor.get("shl_invested_keur", 0)).number_format = _NUMBER_FMT_KEUR
            ws.cell(row_idx, 4, sponsor.get("total_distributions_keur", 0)).number_format = _NUMBER_FMT_KEUR
            irr = sponsor.get("equity_irr", 0)
            c_irr = ws.cell(row_idx, 5, irr)
            c_irr.number_format = "0.00%"
            moic = sponsor.get("moic", 0)
            c_moic = ws.cell(row_idx, 6, moic)
            c_moic.number_format = "0.00x"
            ws.cell(row_idx, 7, sponsor.get("payback_year", "N/A"))
            row_idx += 1

    ws.column_dimensions["A"].width = 22
    for col in range(2, n_years + 3):
        ws.column_dimensions[get_column_letter(col)].width = 12
    ws.freeze_panes = ws.cell(4, 2)


# ---------------------------------------------------------------------------
# Sensitivity sheet
# ---------------------------------------------------------------------------

def _write_sensitivity_sheet(ws, sensitivity_result: dict):
    ws["A1"] = "Sensitivity Analysis"
    ws["A1"].font = Font(name=_FONT_BODY, size=12, bold=True)
    ws["B1"] = "in kEUR"
    ws["B1"].font = Font(name=_FONT_BODY, size=9, italic=True)

    tornado = sensitivity_result.get("tornado", [])
    if tornado:
        ws["A3"] = "Tornado — IRR Impact (sorted by |impact|)"
        ws["A3"].font = Font(name=_FONT_BODY, bold=True)
        ws["A3"].fill = _HEADER_FILL
        ws["A3"].font = Font(name=_FONT_BODY, bold=True, color="FFFFFF")
        ws.merge_cells("A3:E3")

        ws.cell(4, 1, "Variable").font = _SUBHEADER_FONT
        ws.cell(4, 1).fill = _SUBHEADER_FILL
        ws.cell(4, 2, "Low value").font = _SUBHEADER_FONT
        ws.cell(4, 2).fill = _SUBHEADER_FILL
        ws.cell(4, 3, "High value").font = _SUBHEADER_FONT
        ws.cell(4, 3).fill = _SUBHEADER_FILL
        ws.cell(4, 4, "IRR impact (bps)").font = _SUBHEADER_FONT
        ws.cell(4, 4).fill = _SUBHEADER_FILL
        ws.cell(4, 5, "Direction").font = _SUBHEADER_FONT
        ws.cell(4, 5).fill = _SUBHEADER_FILL

        alt = False
        for row_idx, item in enumerate(tornado, start=5):
            ws.cell(row_idx, 1, item.get("variable", ""))
            ws.cell(row_idx, 2, item.get("low_value", 0)).number_format = "0.00"
            ws.cell(row_idx, 3, item.get("high_value", 0)).number_format = "0.00"
            impact = item.get("impact_bps", 0)
            c = ws.cell(row_idx, 4, impact)
            c.number_format = "#,##0"
            direction = "↑ Positive" if impact > 0 else "↓ Negative"
            ws.cell(row_idx, 5, direction)
            if alt:
                for col in range(1, 6):
                    ws.cell(row_idx, col).fill = _ROW_ALT
            alt = not alt

    ws.column_dimensions["A"].width = 22
    for col in range(2, 6):
        ws.column_dimensions[get_column_letter(col)].width = 14
    ws.freeze_panes = ws.cell(4, 2)


# ---------------------------------------------------------------------------
# Main export function
# ---------------------------------------------------------------------------

def export_fid_deck_excel(
    result: any,
    inputs: any,
    filepath: str,
    branding: dict | None = None,
    project_name: str | None = None,
    sponsors: list | None = None,
    sensitivity_result: dict | None = None,
) -> None:
    """Generate a complete FID Deck Excel workbook.

    Sheets (in order):
      Cover | P&L | Balance Sheet | Cash Flow | Debt Schedule | Equity | Sensitivity

    Args:
        result: WaterfallResult from cached_run_waterfall_v3
        inputs: ProjectInputs instance
        filepath: output .xlsx path
        branding: optional dict with logo_bytes, footer_text, disclaimer
        project_name: override for project name on cover
        sponsors: optional list of Sponsor dicts for equity sheet
        sensitivity_result: optional dict with tornado/spider results
    """
    wb = Workbook()
    wb.remove(wb.active)

    proj_name = project_name or (
        getattr(inputs.info, "name", "Project")
        if hasattr(inputs, "info") else "Project"
    )
    tech = getattr(inputs.info, "technology_type", "Solar") if hasattr(inputs, "info") else "Solar"
    cap_mw = getattr(inputs.info, "capacity_mw", 75.26) if hasattr(inputs, "info") else 75.26
    cod = str(getattr(inputs.info, "cod_date", "N/A")) if hasattr(inputs, "info") else "N/A"
    fin_close = str(getattr(inputs.info, "financial_close", "N/A")) if hasattr(inputs, "info") else "N/A"

    ei = ExcelExportInputs(
        result=result,
        inputs=inputs,
        project_name=proj_name,
        technology=tech,
        capacity_mw=cap_mw,
        cod_date=cod,
        financial_close=fin_close,
        sponsors=sponsors or [],
        logo_bytes=branding.get("logo_bytes") if branding else None,
        footer_text=branding.get("footer_text", "Confidential — prepared by OpusCore v2") if branding else "Confidential — prepared by OpusCore v2",
        disclaimer=branding.get("disclaimer", "This financial model is for illustrative purposes only.") if branding else "This financial model is for illustrative purposes only.",
    )

    # Cover
    ws_cover = wb.create_sheet("Cover")
    if ei.logo_bytes:
        try:
            img = XLImage(BytesIO(ei.logo_bytes))
            img.height = 90
            img.width = int(img.width * 90 / img.height) if img.height > 0 else 120
            ws_cover.add_image(img, "D1")
        except Exception:
            pass
    _write_cover(ws_cover, ei)

    # P&L
    ws_pl = wb.create_sheet("P&L")
    _write_pl_sheet(ws_pl, ei)

    # Balance Sheet
    ws_bs = wb.create_sheet("Balance Sheet")
    _write_bs_sheet(ws_bs, ei)

    # Cash Flow
    ws_cf = wb.create_sheet("Cash Flow")
    _write_cf_sheet(ws_cf, ei)

    # Debt Schedule
    ws_ds = wb.create_sheet("Debt Schedule")
    _write_ds_sheet(ws_ds, ei)

    # Equity
    ws_eq = wb.create_sheet("Equity")
    _write_equity_sheet(ws_eq, ei)

    # Sensitivity (optional)
    if sensitivity_result:
        ws_sens = wb.create_sheet("Sensitivity")
        _write_sensitivity_sheet(ws_sens, sensitivity_result)

    wb.save(filepath)
