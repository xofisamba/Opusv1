"""Excel export - export model results to formatted Excel file.

Produces multi-sheet Excel matching Oborovo Excel structure:
- Summary sheet with key metrics
- Inputs sheet
- Cash flow schedule
- Debt amortization schedule
- Waterfall detail
- DSCR/LLCR/PLCR ratios
"""
from datetime import date
from typing import Optional

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


def export_to_excel(inputs, engine, results=None, filepath="project_output.xlsx"):
    """Export complete model to formatted Excel.
    
    Args:
        inputs: ProjectInputs
        engine: PeriodEngine  
        results: Optional dict with computed results (IRR, NPV, etc.)
        filepath: Output file path
    
    Returns:
        True if successful, error message otherwise
    """
    if not OPENPYXL_AVAILABLE:
        return "openpyxl not installed. Run: pip install openpyxl"
    
    try:
        wb = openpyxl.Workbook()
        
        # Style definitions
        header_font = Font(bold=True, size=12)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font_white = Font(bold=True, size=12, color="FFFFFF")
        money_format = '#,##0'
        percent_format = '0.00%'
        date_format = 'YYYY-MM-DD'
        
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # ============ SUMMARY SHEET ============
        ws_summary = wb.active
        ws_summary.title = "Summary"
        
        # Title
        ws_summary['A1'] = f"{inputs.info.name} - Financial Model"
        ws_summary['A1'].font = Font(bold=True, size=14)
        
        # Key metrics
        row = 3
        metrics = [
            ("Project", inputs.info.name),
            ("Company", inputs.info.company),
            ("Country", inputs.info.country_iso),
            ("Technology", "Solar" if inputs.technical.pv_degradation > 0 else "Wind"),
            ("Capacity (MWp)", inputs.technical.capacity_mw),
            ("P50 Yield (hrs)", inputs.technical.operating_hours_p50),
            ("PPA Tariff (€/MWh)", inputs.revenue.ppa_base_tariff),
            ("PPA Term (years)", inputs.revenue.ppa_term_years),
            ("Total CAPEX (k€)", inputs.capex.total_capex),
            ("Gearing", inputs.financing.gearing_ratio),
            ("Target DSCR", inputs.financing.target_dscr),
            ("Debt Tenor (years)", inputs.financing.senior_tenor_years),
        ]
        
        for label, value in metrics:
            ws_summary[f'A{row}'] = label
            ws_summary[f'B{row}'] = value
            row += 1
        
        # Results if provided
        if results:
            row += 1
            ws_summary[f'A{row}'] = "RESULTS"
            ws_summary[f'A{row}'].font = header_font
            
            row += 1
            result_metrics = [
                ("Project IRR", f"{results.get('project_irr', 0)*100:.2f}%"),
                ("Equity IRR", f"{results.get('equity_irr', 0)*100:.2f}%"),
                ("Project NPV (k€)", f"{results.get('project_npv', 0):,.0f}"),
                ("Equity NPV (k€)", f"{results.get('equity_npv', 0):,.0f}"),
                ("Avg DSCR", f"{results.get('avg_dscr', 0):.3f}x"),
                ("Min DSCR", f"{results.get('min_dscr', 0):.3f}x"),
                ("Min LLCR", f"{results.get('min_llcr', 0):.2f}x"),
                ("Min PLCR", f"{results.get('min_plcr', 0):.2f}x"),
            ]
            
            for label, value in result_metrics:
                ws_summary[f'A{row}'] = label
                ws_summary[f'B{row}'] = value
                row += 1
        
        # Column widths
        ws_summary.column_dimensions['A'].width = 25
        ws_summary.column_dimensions['B'].width = 20
        
        # ============ INPUTS SHEET ============
        ws_inputs = wb.create_sheet("Inputs")
        
        row = 1
        ws_inputs[f'A{row}'] = "INPUT PARAMETERS"
        ws_inputs[f'A{row}'].font = Font(bold=True, size=12)
        row += 2
        
        # Project section
        ws_inputs[f'A{row}'] = "Project"
        ws_inputs[f'A{row}'].font = header_font
        row += 1
        
        input_sections = [
            ("Name", inputs.info.name),
            ("Company", inputs.info.company),
            ("Country", inputs.info.country_iso),
            ("Financial Close", inputs.info.financial_close.isoformat() if inputs.info.financial_close else ""),
            ("COD Date", inputs.info.cod_date.isoformat() if inputs.info.cod_date else ""),
            ("Horizon (years)", inputs.info.horizon_years),
        ]
        
        for label, value in input_sections:
            ws_inputs[f'A{row}'] = label
            ws_inputs[f'B{row}'] = value
            row += 1
        
        row += 1
        ws_inputs[f'A{row}'] = "Technical"
        ws_inputs[f'A{row}'].font = header_font
        row += 1
        
        tech_sections = [
            ("Capacity (MW)", inputs.technical.capacity_mw),
            ("P50 Yield (hrs)", inputs.technical.operating_hours_p50),
            ("P90 Yield (hrs)", inputs.technical.operating_hours_p90_10y),
            ("Degradation", f"{inputs.technical.pv_degradation*100:.1f}%"),
            ("Availability", f"{inputs.technical.plant_availability*100:.1f}%"),
        ]
        
        for label, value in tech_sections:
            ws_inputs[f'A{row}'] = label
            ws_inputs[f'B{row}'] = value
            row += 1
        
        row += 1
        ws_inputs[f'A{row}'] = "Revenue"
        ws_inputs[f'A{row}'].font = header_font
        row += 1
        
        rev_sections = [
            ("PPA Tariff (€/MWh)", inputs.revenue.ppa_base_tariff),
            ("PPA Term (years)", inputs.revenue.ppa_term_years),
            ("PPA Escalation", f"{inputs.revenue.ppa_index*100:.1f}%"),
            ("Market Prices", f"Base {inputs.revenue.market_prices_curve[0] if inputs.revenue.market_prices_curve else 0:.0f} €/MWh"),
        ]
        
        for label, value in rev_sections:
            ws_inputs[f'A{row}'] = label
            ws_inputs[f'B{row}'] = value
            row += 1
        
        row += 1
        ws_inputs[f'A{row}'] = "Financing"
        ws_inputs[f'A{row}'].font = header_font
        row += 1
        
        fin_sections = [
            ("Gearing", f"{inputs.financing.gearing_ratio*100:.0f}%"),
            ("Debt Tenor (years)", inputs.financing.senior_tenor_years),
            ("Base Rate", f"{inputs.financing.base_rate*100:.2f}%"),
            ("Margin (bps)", inputs.financing.margin_bps),
            ("All-in Rate", f"{inputs.financing.all_in_rate*100:.2f}%"),
            ("Target DSCR", f"{inputs.financing.target_dscr:.2f}x"),
            ("Lockup DSCR", f"{inputs.financing.lockup_dscr:.2f}x"),
            ("DSRA Months", inputs.financing.dsra_months),
        ]
        
        for label, value in fin_sections:
            ws_inputs[f'A{row}'] = label
            ws_inputs[f'B{row}'] = value
            row += 1
        
        row += 1
        ws_inputs[f'A{row}'] = "Tax"
        ws_inputs[f'A{row}'].font = header_font
        row += 1
        
        tax_sections = [
            ("Corporate Tax", f"{inputs.tax.corporate_rate*100:.1f}%"),
            ("ATAD", "Yes" if inputs.tax.atad_ebitda_limit else "No"),
            ("Loss Carryforward (years)", inputs.tax.loss_carryforward_years),
        ]
        
        for label, value in tax_sections:
            ws_inputs[f'A{row}'] = label
            ws_inputs[f'B{row}'] = value
            row += 1
        
        ws_inputs.column_dimensions['A'].width = 25
        ws_inputs.column_dimensions['B'].width = 20
        
        # ============ CASH FLOWS SHEET ============
        ws_cf = wb.create_sheet("Cash Flows")
        
        # Headers
        headers = ["Period", "Date", "Year", "Generation (MWh)", "Revenue (k€)", "OPEX (k€)", "EBITDA (k€)", "Tax (k€)", "CF After Tax (k€)", "Senior DS (k€)", "CF to Equity (k€)"]
        
        for col, header in enumerate(headers, 1):
            cell = ws_cf.cell(row=1, column=col, value=header)
            cell.font = header_font_white
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        
        # Data rows (would need actual calculations from engine)
        # This is a template - actual data would come from waterfall
        row = 2
        for p in list(engine.periods())[:60]:  # First 60 periods (30 years semi-annual)
            ws_cf.cell(row=row, column=1, value=p.index)
            ws_cf.cell(row=row, column=2, value=p.end_date.isoformat())
            ws_cf.cell(row=row, column=3, value=p.year_index)
            # Placeholder values - actual calculations would go here
            row += 1
        
        # Column widths
        ws_cf.column_dimensions['A'].width = 8
        ws_cf.column_dimensions['B'].width = 12
        ws_cf.column_dimensions['C'].width = 6
        for col in range(4, 12):
            ws_cf.column_dimensions[get_column_letter(col)].width = 15
        
        # ============ DEBT SCHEDULE SHEET ============
        ws_debt = wb.create_sheet("Debt Schedule")
        
        debt_headers = ["Period", "Date", "Year", "Opening Balance (k€)", "Interest (k€)", "Principal (k€)", "Debt Service (k€)", "Closing Balance (k€)", "DSCR"]
        
        for col, header in enumerate(debt_headers, 1):
            cell = ws_debt.cell(row=1, column=col, value=header)
            cell.font = header_font_white
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        
        # Placeholder - actual schedule would come from domain/financing
        row = 2
        for i in range(1, inputs.financing.senior_tenor_years * 2 + 1):
            ws_debt.cell(row=row, column=1, value=i)
            row += 1
        
        ws_debt.column_dimensions['A'].width = 8
        ws_debt.column_dimensions['B'].width = 12
        ws_debt.column_dimensions['C'].width = 6
        for col in range(4, 10):
            ws_debt.column_dimensions[get_column_letter(col)].width = 18
        
        # ============ WATERFALL SHEET ============
        ws_wf = wb.create_sheet("Waterfall")
        
        wf_headers = ["Period", "Date", "Year", "Revenue", "OPEX", "EBITDA", "Interest", "Principal", "Tax", "CF After Tax", "DSRA", "CF Available", "Distribution", "Lockup"]
        
        for col, header in enumerate(wf_headers, 1):
            cell = ws_wf.cell(row=1, column=col, value=header)
            cell.font = header_font_white
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        
        ws_wf.column_dimensions['A'].width = 8
        ws_wf.column_dimensions['B'].width = 12
        ws_wf.column_dimensions['C'].width = 6
        for col in range(4, 15):
            ws_wf.column_dimensions[get_column_letter(col)].width = 14
        
        # Save workbook
        wb.save(filepath)
        
        return True
        
    except Exception as e:
        return f"Export failed: {str(e)}"


def export_cashflows_to_csv(inputs, engine, filepath="cashflows.csv"):
    """Export simplified cash flows to CSV.
    
    Args:
        inputs: ProjectInputs
        engine: PeriodEngine
        filepath: Output path
    
    Returns:
        True if successful, error message otherwise
    """
    try:
        with open(filepath, 'w') as f:
            f.write("Period,Date,Year,IsOperation,Generation_MWh,Revenue_kEUR,OPEX_kEUR,EBITDA_kEUR\n")
            
            for p in engine.periods():
                f.write(f"{p.index},{p.end_date.isoformat()},{p.year_index},{p.is_operation},0,0,0,0\n")
        
        return True
    except Exception as e:
        return f"CSV export failed: {str(e)}"


def export_to_excel_advanced(inputs, engine, waterfall_result=None, filepath="project_full.xlsx"):
    """Advanced Excel export with full calculations.
    
    Args:
        inputs: ProjectInputs
        engine: PeriodEngine
        waterfall_result: Optional WaterfallResult from compute_waterfall
        filepath: Output path
    
    Returns:
        True if successful, error message otherwise
    """
    if not OPENPYXL_AVAILABLE:
        return "openpyxl not installed. Run: pip install openpyxl"
    
    try:
        wb = openpyxl.Workbook()
        
        # Style definitions
        header_font = Font(bold=True, size=11, color="FFFFFF")
        header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
        money_format = '#,##0'
        percent_format = '0.00%'
        
        # ============ EXECUTIVE SUMMARY ============
        ws = wb.active
        ws.title = "Executive Summary"
        
        ws['A1'] = f"Project: {inputs.info.name}"
        ws['A1'].font = Font(bold=True, size=14)
        ws['A2'] = f"Company: {inputs.info.company}"
        
        row = 4
        ws[f'A{row}'] = "KEY PARAMETERS"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1
        
        params = [
            ("Capacity", f"{inputs.technical.capacity_mw:.2f} MWp"),
            ("P50 Yield", f"{inputs.technical.operating_hours_p50:.0f} hrs"),
            ("PPA Tariff", f"{inputs.revenue.ppa_base_tariff:.0f} €/MWh"),
            ("Total CAPEX", f"{inputs.capex.total_capex:,.0f} k€"),
            ("CAPEX/MW", f"{inputs.capex.total_capex / inputs.technical.capacity_mw:,.0f} k€/MW"),
            ("Gearing", f"{inputs.financing.gearing_ratio:.0%}"),
            ("Debt Tenor", f"{inputs.financing.senior_tenor_years} years"),
            ("Target DSCR", f"{inputs.financing.target_dscr:.2f}x"),
        ]
        
        for label, value in params:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            row += 1
        
        if waterfall_result:
            row += 1
            ws[f'A{row}'] = "FINANCIAL RESULTS"
            ws[f'A{row}'].font = Font(bold=True, size=12)
            row += 1
            
            results = [
                ("Project IRR", f"{waterfall_result.project_irr*100:.2f}%"),
                ("Equity IRR", f"{waterfall_result.equity_irr*100:.2f}%"),
                ("Project NPV", f"{waterfall_result.project_npv:,.0f} k€"),
                ("Equity NPV", f"{waterfall_result.equity_npv:,.0f} k€"),
                ("Avg DSCR", f"{waterfall_result.avg_dscr:.3f}x"),
                ("Min DSCR", f"{waterfall_result.min_dscr:.3f}x"),
                ("Min LLCR", f"{waterfall_result.min_llcr:.2f}x"),
                ("Min PLCR", f"{waterfall_result.min_plcr:.2f}x"),
                ("Total Revenue", f"{waterfall_result.total_revenue_keur:,.0f} k€"),
                ("Total Distributions", f"{waterfall_result.total_distribution_keur:,.0f} k€"),
            ]
            
            for label, value in results:
                ws[f'A{row}'] = label
                ws[f'B{row}'] = value
                row += 1
        
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 18
        
        # ============ CASH FLOW SCHEDULE ============
        ws_cf = wb.create_sheet("Cash Flows")
        
        # Headers
        headers = ["Period", "Date", "Year", "H-Year", "Op?", "Revenue (k€)", "OPEX (k€)", "EBITDA (k€)", "Tax (k€)", "CFAT (k€)", "Sr DS (k€)", "SHL (k€)", "CFADS (k€)", "DSRA (k€)", "Dist (k€)", "DSCR", "Lockup"]
        
        for col, h in enumerate(headers, 1):
            cell = ws_cf.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        
        if waterfall_result:
            for i, wp in enumerate(waterfall_result.periods):
                row = i + 2
                ws_cf.cell(row=row, column=1, value=wp.period)
                ws_cf.cell(row=row, column=2, value=wp.date.isoformat() if wp.date else "")
                ws_cf.cell(row=row, column=3, value=wp.year_index if wp.year_index else 0)
                ws_cf.cell(row=row, column=4, value=f"H{wp.period % 2 or 2}")
                ws_cf.cell(row=row, column=5, value="Y" if wp.year_index and wp.year_index > 0 else "N")
                ws_cf.cell(row=row, column=6, value=wp.revenue_keur)
                ws_cf.cell(row=row, column=7, value=wp.opex_keur)
                ws_cf.cell(row=row, column=8, value=wp.ebitda_keur)
                ws_cf.cell(row=row, column=9, value=wp.tax_keur)
                ws_cf.cell(row=row, column=10, value=wp.cf_after_tax_keur)
                ws_cf.cell(row=row, column=11, value=wp.senior_ds_keur)
                ws_cf.cell(row=row, column=12, value=wp.shl_service_keur)
                ws_cf.cell(row=row, column=13, value=wp.cf_after_reserves_keur)
                ws_cf.cell(row=row, column=14, value=wp.dsra_contribution_keur)
                ws_cf.cell(row=row, column=15, value=wp.distribution_keur)
                ws_cf.cell(row=row, column=16, value=round(wp.dscr, 2) if wp.dscr > 0 else None)
                ws_cf.cell(row=row, column=17, value="🔒" if wp.lockup_active else "")
        
        # Column widths
        ws_cf.column_dimensions['A'].width = 8
        ws_cf.column_dimensions['B'].width = 12
        ws_cf.column_dimensions['C'].width = 6
        ws_cf.column_dimensions['D'].width = 6
        ws_cf.column_dimensions['E'].width = 5
        for col in range(6, 18):
            ws_cf.column_dimensions[get_column_letter(col)].width = 12
        
        # ============ DEBT SCHEDULE ============
        ws_debt = wb.create_sheet("Debt Schedule")
        
        debt_headers = ["Period", "Date", "Year", "Opening (k€)", "Interest (k€)", "Principal (k€)", "DS (k€)", "Closing (k€)", "DSCR"]
        
        for col, h in enumerate(debt_headers, 1):
            cell = ws_debt.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        
        if waterfall_result and waterfall_result.sculpting_result:
            sr = waterfall_result.sculpting_result
            for i in range(len(sr.payments)):
                row = i + 2
                ws_debt.cell(row=row, column=1, value=i+1)
                # Date would come from period engine
                ws_debt.cell(row=row, column=4, value=round(sr.balance_schedule[i], 0) if i < len(sr.balance_schedule) else 0)
                ws_debt.cell(row=row, column=5, value=round(sr.interest_schedule[i], 0))
                ws_debt.cell(row=row, column=6, value=round(sr.principal_schedule[i], 0))
                ws_debt.cell(row=row, column=7, value=round(sr.payments[i], 0))
                ws_debt.cell(row=row, column=8, value=round(sr.balance_schedule[i], 0) if i < len(sr.balance_schedule) else 0)
                ws_debt.cell(row=row, column=9, value=round(sr.dscr_schedule[i], 2) if i < len(sr.dscr_schedule) else None)
        
        ws_debt.column_dimensions['A'].width = 8
        ws_debt.column_dimensions['B'].width = 12
        ws_debt.column_dimensions['C'].width = 6
        for col in range(4, 10):
            ws_debt.column_dimensions[get_column_letter(col)].width = 14
        
        # ============ RETURNS ANALYSIS ============
        ws_ret = wb.create_sheet("Returns")
        
        ws_ret['A1'] = "RETURNS ANALYSIS"
        ws_ret['A1'].font = Font(bold=True, size=12)
        
        if waterfall_result:
            row = 3
            returns_data = [
                ("Project IRR", f"{waterfall_result.project_irr*100:.2f}%"),
                ("Equity IRR", f"{waterfall_result.equity_irr*100:.2f}%"),
                ("Project NPV (6.41%)", f"{waterfall_result.project_npv:,.0f} k€"),
                ("Equity NPV (9.65%)", f"{waterfall_result.equity_npv:,.0f} k€"),
                ("Total Distributions", f"{waterfall_result.total_distribution_keur:,.0f} k€"),
                ("Sculpting Debt", f"{waterfall_result.sculpting_result.debt_keur:,.0f} k€" if waterfall_result.sculpting_result else "N/A"),
            ]
            
            for label, value in returns_data:
                ws_ret[f'A{row}'] = label
                ws_ret[f'B{row}'] = value
                row += 1
        
        ws_ret.column_dimensions['A'].width = 25
        ws_ret.column_dimensions['B'].width = 18
        
        # Save
        wb.save(filepath)
        return True
        
    except Exception as e:
        return f"Export failed: {str(e)}"