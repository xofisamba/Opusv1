"""Excel parser for Oborovo-style project finance models.

Reads an Oborovo .xlsm file and produces a fully-populated ProjectInputs object.
This is a one-time conversion - the Python model uses the parsed inputs,
not the Excel file.

Supports:
- Oborovo Excel (17 sheets, semi-annual periods)
- Standard project finance Excel exports from similar templates

Example:
    >>> inputs = parse_oborovo_excel("path/to/Oborovo.xlsm")
    >>> print(inputs.capex.total_capex)
    56899.5
"""
from dataclasses import dataclass
from datetime import date
from typing import Any, Optional
import json

try:
    from openpyxl import load_workbook
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


@dataclass
class ExcelParserConfig:
    """Configuration for Excel parser behavior."""
    data_only: bool = True  # Read cached values, not formulas
    check_invalid: bool = True  # Validate input ranges
    warn_on_missing: bool = True  # Log warnings for missing cells


class ExcelParserError(Exception):
    """Raised when Excel parsing fails."""
    pass


def _safe_get(ws, cell_ref: str, default: Any = None) -> Any:
    """Safely get a cell value from worksheet.
    
    Args:
        ws: openpyxl worksheet
        cell_ref: Cell reference like 'D2'
        default: Default value if cell is None or empty
    
    Returns:
        Cell value or default
    """
    try:
        val = ws[cell_ref].value
        if val is None or val == "":
            return default
        return val
    except Exception:
        return default


def _safe_date(ws, cell_ref: str) -> Optional[date]:
    """Safely parse a date cell.
    
    Args:
        ws: openpyxl worksheet
        cell_ref: Cell reference
    
    Returns:
        date object or None
    """
    val = _safe_get(ws, cell_ref)
    if val is None:
        return None
    if isinstance(val, date):
        return val
    if isinstance(val, str):
        # Try parsing common formats
        for fmt in ["%Y-%m-%d", "%d.%m.%Y", "%d/%m/%Y"]:
            try:
                return date.fromisoformat(val.replace("/", "-"))
            except ValueError:
                continue
    # Assume it's a datetime object from openpyxl
    try:
        return val.date()
    except AttributeError:
        return None


def _safe_float(val: Any, default: float = 0.0) -> float:
    """Safely convert to float."""
    if val is None:
        return default
    try:
        return float(val)
    except (TypeError, ValueError):
        return default


def _safe_int(val: Any, default: int = 0) -> int:
    """Safely convert to int."""
    if val is None:
        return default
    try:
        return int(val)
    except (TypeError, ValueError):
        return default


def parse_project_info(ws) -> dict:
    """Parse ProjectInfo from Inputs sheet.
    
    Corresponds to Excel Inputs rows 2-18.
    """
    return {
        "name": _safe_get(ws, "D2", "Unknown Project"),
        "company": _safe_get(ws, "D3", "Unknown Company"),
        "code": _safe_get(ws, "D4", "UNK"),
        "country_iso": str(_safe_get(ws, "D5", "XX"))[:2].upper(),
        "financial_close": _safe_date(ws, "D9"),
        "construction_months": _safe_int(_safe_get(ws, "D10", 12)),
        "cod_date": _safe_date(ws, "D11"),
        "horizon_years": _safe_int(_safe_get(ws, "D14", 30)),
        "period_frequency": _safe_get(ws, "D18", "Semestrial"),
    }


def parse_technical_params(ws) -> dict:
    """Parse TechnicalParams from Inputs sheet.
    
    Corresponds to Excel Inputs rows 51-68.
    """
    return {
        "capacity_mw": _safe_float(_safe_get(ws, "D51", 0)),
        "yield_scenario": _safe_get(ws, "D52", "P_50"),
        "operating_hours_p50": _safe_float(_safe_get(ws, "D64", 1494)),
        "operating_hours_p90_10y": _safe_float(_safe_get(ws, "D68", 1410)),
        "pv_degradation": _safe_float(_safe_get(ws, "D56", 0.004)),
        "bess_degradation": _safe_float(_safe_get(ws, "D57", 0.003)),
        "plant_availability": _safe_float(_safe_get(ws, "D58", 0.99)),
        "grid_availability": _safe_float(_safe_get(ws, "D59", 0.99)),
        "bess_enabled": bool(_safe_get(ws, "D140", False)),
    }


def parse_financing_params(ws) -> dict:
    """Parse FinancingParams from Inputs sheet.
    
    Corresponds to Excel Inputs rows 168-349.
    """
    return {
        # Equity structure
        "share_capital_keur": _safe_float(_safe_get(ws, "D312", 500)),
        "share_premium_keur": _safe_float(_safe_get(ws, "D313", 0)),
        "shl_amount_keur": _safe_float(_safe_get(ws, "D325", 0)),
        "shl_rate": _safe_float(_safe_get(ws, "F328", 0.08)),
        # Debt structure
        "gearing_ratio": _safe_float(_safe_get(ws, "D168", 0.75)),
        "senior_tenor_years": _safe_int(_safe_get(ws, "D196", 14)),
        "base_rate": _safe_float(_safe_get(ws, "D202", 0.03)),
        "margin_bps": _safe_int(_safe_get(ws, "D203", 265)),
        "floating_share": _safe_float(_safe_get(ws, "B39", 0.2)),
        "fixed_share": _safe_float(_safe_get(ws, "B40", 0.8)),
        "hedge_coverage": _safe_float(_safe_get(ws, "D230", 0.8)),
        # Fees
        "commitment_fee": _safe_float(_safe_get(ws, "D214", 0.0105)),
        "arrangement_fee": _safe_float(_safe_get(ws, "D218", 0)),
        "structuring_fee": _safe_float(_safe_get(ws, "D217", 0.01)),
        # Covenants
        "target_dscr": _safe_float(_safe_get(ws, "D221", 1.15)),
        "lockup_dscr": _safe_float(_safe_get(ws, "D223", 1.10)),
        "min_llcr": _safe_float(_safe_get(ws, "D224", 1.15)),
        # Reserves
        "dsra_months": _safe_int(_safe_get(ws, "D348", 6)),
    }


def parse_tax_params(ws) -> dict:
    """Parse TaxParams from Inputs sheet.
    
    Corresponds to Excel Inputs rows 403-426.
    """
    return {
        "corporate_rate": _safe_float(_safe_get(ws, "D403", 0.10)),
        "loss_carryforward_years": _safe_int(_safe_get(ws, "D407", 5)),
        "loss_carryforward_cap": _safe_float(_safe_get(ws, "D408", 1.0)),
        "legal_reserve_cap": _safe_float(_safe_get(ws, "D410", 0.10)),
        "thin_cap_enabled": bool(_safe_get(ws, "D414", False)),
        "thin_cap_de_ratio": _safe_float(_safe_get(ws, "D415", 0.8)),
        "atad_ebitda_limit": _safe_float(_safe_get(ws, "D416", 0.30)),
        "atad_min_interest_keur": _safe_float(_safe_get(ws, "D417", 3000)),
        "wht_sponsor_dividends": _safe_float(_safe_get(ws, "D426", 0.05)),
        "wht_sponsor_shl_interest": _safe_float(_safe_get(ws, "D423", 0)),
        "shl_cap_applies": bool(_safe_get(ws, "D412", True)),
    }


def parse_revenue_params(ws) -> dict:
    """Parse RevenueParams from Inputs sheet.
    
    Corresponds to Excel Inputs rows 78-141.
    """
    # Market prices from row 107 (columns E-AJ for years 1-30)
    market_prices = []
    for col in range(5, 35):  # E=5 to AJ=34
        val = _safe_float(ws.cell(row=107, column=col).value, 0)
        if val > 0:
            market_prices.append(val)
    
    return {
        "ppa_base_tariff": _safe_float(_safe_get(ws, "D78", 57)),
        "ppa_term_years": _safe_int(_safe_get(ws, "D81", 12)),
        "ppa_index": _safe_float(_safe_get(ws, "D83", 0.02)),
        "ppa_production_share": _safe_float(_safe_get(ws, "D80", 1.0)),
        "market_scenario": _safe_get(ws, "B103", "Central"),
        "market_prices_curve": tuple(market_prices) if market_prices else (),
        "market_inflation": _safe_float(_safe_get(ws, "B129", 0.02)),
        "balancing_cost_pv": _safe_float(_safe_get(ws, "D114", 0.025)),
        "balancing_cost_bess": _safe_float(_safe_get(ws, "D115", 0.025)),
        "co2_enabled": bool(_safe_get(ws, "D139", False)),
        "co2_price_eur": _safe_float(_safe_get(ws, "E141", 1.5)),
    }


def parse_opex_items(ws) -> list:
    """Parse OPEX items from Inputs sheet.
    
    Corresponds to Excel Inputs rows 146-161 (15 OPEX categories).
    """
    opex_items = []
    
    # Row 146 header, rows 147-161 contain item data
    for row in range(147, 162):
        name = _safe_get(ws, f"C{row}", f"OPEX_{row-146}")
        y1_amount = _safe_float(_safe_get(ws, f"D{row}", 0))
        inflation = _safe_float(_safe_get(ws, f"E{row}", 0.02))
        
        if y1_amount > 0:
            opex_items.append({
                "name": name,
                "y1_amount_keur": y1_amount,
                "annual_inflation": inflation,
            })
    
    return opex_items


def parse_capex_items(ws) -> dict:
    """Parse CAPEX items from Inputs sheet.
    
    Corresponds to Excel Inputs rows 23-44 (22 CAPEX categories).
    Returns a dict mapping item names to their values.
    """
    items = {}
    
    # Rows 23-44 contain CAPEX items
    for row in range(23, 45):
        name = _safe_get(ws, f"B{row}", f"CAPEX_{row-22}")
        amount = _safe_float(_safe_get(ws, f"C{row}", 0))
        y0_share = _safe_float(_safe_get(ws, f"D{row}", 0))
        
        # Spending profile in columns E-H (Y1-Y4)
        profile = []
        for col in ["E", "F", "G", "H"]:
            profile.append(_safe_float(_safe_get(ws, f"{col}{row}", 0)))
        
        if amount > 0:
            items[name] = {
                "amount_keur": amount,
                "y0_share": y0_share,
                "spending_profile": tuple(profile),
            }
    
    return items


def create_project_inputs_from_excel_dict(data: dict) -> "ProjectInputs":
    """Create ProjectInputs from parsed Excel data dictionary.
    
    Args:
        data: Dictionary with keys matching parse_* function outputs
    
    Returns:
        ProjectInputs instance
    """
    from domain.inputs import (
        ProjectInputs, ProjectInfo, CapexStructure, CapexItem,
        OpexItem, TechnicalParams, RevenueParams, FinancingParams, TaxParams,
        PeriodFrequency,
    )
    
    info = ProjectInfo(
        name=data["info"]["name"],
        company=data["info"]["company"],
        code=data["info"]["code"],
        country_iso=data["info"]["country_iso"],
        financial_close=data["info"]["financial_close"],
        construction_months=data["info"]["construction_months"],
        cod_date=data["info"]["cod_date"],
        horizon_years=data["info"]["horizon_years"],
        period_frequency=PeriodFrequency(data["info"]["period_frequency"]),
    )
    
    technical = TechnicalParams(**data["technical"])
    revenue = RevenueParams(**data["revenue"])
    financing = FinancingParams(**data["financing"])
    tax = TaxParams(**data["tax"])
    
    # Build CAPEX structure
    capex_items = data.get("capex", {})
    capex_kwargs = {}
    
    # Map common names to CapexStructure fields
    name_to_field = {
        "EPC Contract": "epc_contract",
        "Production Units": "production_units",
        "Other EPC": "epc_other",
        "Grid Connection": "grid_connection",
        "Operations Preparation": "ops_prep",
        "Insurances": "insurances",
        "Lease & Property Tax": "lease_tax",
        "Construction Management A": "construction_mgmt_a",
        "Commissioning": "commissioning",
        "Audit & Legal": "audit_legal",
        "Construction Management B": "construction_mgmt_b",
        "Contingencies": "contingencies",
        "Taxes & Duties": "taxes",
        "Project Acquisition": "project_acquisition",
        "Project Rights": "project_rights",
    }
    
    for name, field in name_to_field.items():
        item_data = capex_items.get(name, {"amount_keur": 0, "y0_share": 0, "spending_profile": ()})
        capex_kwargs[field] = CapexItem(
            name=name,
            amount_keur=item_data["amount_keur"],
            y0_share=item_data["y0_share"],
            spending_profile=item_data["spending_profile"],
        )
    
    # Add dynamic items from Excel
    capex_kwargs["idc_keur"] = data.get("capex", {}).get("IDC", {}).get("amount_keur", 0)
    capex_kwargs["commitment_fees_keur"] = data.get("capex", {}).get("Commitment Fees", {}).get("amount_keur", 0)
    capex_kwargs["bank_fees_keur"] = data.get("capex", {}).get("Bank Fees", {}).get("amount_keur", 0)
    capex_kwargs["vat_costs_keur"] = data.get("capex", {}).get("VAT Costs", {}).get("amount_keur", 0)
    capex_kwargs["reserve_accounts_keur"] = data.get("capex", {}).get("Reserve Accounts", {}).get("amount_keur", 0)
    
    capex = CapexStructure(**capex_kwargs)
    
    # Build OPEX items
    opex_items = tuple(
        OpexItem(**item) for item in data.get("opex", [])
    )
    
    return ProjectInputs(
        info=info,
        technical=technical,
        capex=capex,
        opex=opex_items,
        revenue=revenue,
        financing=financing,
        tax=tax,
    )


def parse_oborovo_excel(path: str, config: Optional[ExcelParserConfig] = None) -> "ProjectInputs":
    """Parse Oborovo Excel file into ProjectInputs.
    
    Args:
        path: Path to .xlsm file
        config: Optional parser configuration
    
    Returns:
        ProjectInputs with all values from Excel
    
    Raises:
        ExcelParserError: If file cannot be parsed
    """
    if not OPENPYXL_AVAILABLE:
        raise ExcelParserError("openpyxl is required for Excel parsing. Install with: pip install openpyxl")
    
    if config is None:
        config = ExcelParserConfig()
    
    try:
        wb = load_workbook(path, data_only=config.data_only)
    except FileNotFoundError:
        raise ExcelParserError(f"File not found: {path}")
    except Exception as e:
        raise ExcelParserError(f"Cannot open Excel file: {e}")
    
    if "Inputs" not in wb.sheetnames:
        raise ExcelParserError(f"'Inputs' sheet not found in {path}")
    
    ws = wb["Inputs"]
    
    # Parse all sections
    data = {
        "info": parse_project_info(ws),
        "technical": parse_technical_params(ws),
        "capex": parse_capex_items(ws),
        "opex": parse_opex_items(ws),
        "revenue": parse_revenue_params(ws),
        "financing": parse_financing_params(ws),
        "tax": parse_tax_params(ws),
    }
    
    # Validate required fields
    if config.check_invalid:
        errors = []
        if data["info"]["horizon_years"] <= 0:
            errors.append("Invalid horizon_years")
        if data["technical"]["capacity_mw"] <= 0:
            errors.append("Invalid capacity_mw")
        if data["financing"]["gearing_ratio"] <= 0 or data["financing"]["gearing_ratio"] > 1:
            errors.append("Invalid gearing_ratio")
        
        if errors:
            raise ExcelParserError(f"Validation errors: {', '.join(errors)}")
    
    return create_project_inputs_from_excel_dict(data)


def export_inputs_to_json(inputs: "ProjectInputs", path: str) -> None:
    """Export ProjectInputs to JSON for caching/reuse.
    
    Args:
        inputs: ProjectInputs instance
        path: Output JSON file path
    """
    # Convert to dict (simple serialization)
    data = {
        "info": {
            "name": inputs.info.name,
            "company": inputs.info.company,
            "code": inputs.info.code,
            "country_iso": inputs.info.country_iso,
            "financial_close": inputs.info.financial_close.isoformat(),
            "construction_months": inputs.info.construction_months,
            "cod_date": inputs.info.cod_date.isoformat(),
            "horizon_years": inputs.info.horizon_years,
            "period_frequency": inputs.info.period_frequency.value,
        },
        "technical": {
            "capacity_mw": inputs.technical.capacity_mw,
            "yield_scenario": inputs.technical.yield_scenario,
            "operating_hours_p50": inputs.technical.operating_hours_p50,
            "operating_hours_p90_10y": inputs.technical.operating_hours_p90_10y,
            "pv_degradation": inputs.technical.pv_degradation,
            "bess_degradation": inputs.technical.bess_degradation,
            "plant_availability": inputs.technical.plant_availability,
            "grid_availability": inputs.technical.grid_availability,
            "bess_enabled": inputs.technical.bess_enabled,
        },
        "revenue": {
            "ppa_base_tariff": inputs.revenue.ppa_base_tariff,
            "ppa_term_years": inputs.revenue.ppa_term_years,
            "ppa_index": inputs.revenue.ppa_index,
            "ppa_production_share": inputs.revenue.ppa_production_share,
            "market_scenario": inputs.revenue.market_scenario,
            "market_prices_curve": list(inputs.revenue.market_prices_curve),
            "market_inflation": inputs.revenue.market_inflation,
            "balancing_cost_pv": inputs.revenue.balancing_cost_pv,
            "balancing_cost_bess": inputs.revenue.balancing_cost_bess,
            "co2_enabled": inputs.revenue.co2_enabled,
            "co2_price_eur": inputs.revenue.co2_price_eur,
        },
        "financing": {
            "share_capital_keur": inputs.financing.share_capital_keur,
            "share_premium_keur": inputs.financing.share_premium_keur,
            "shl_amount_keur": inputs.financing.shl_amount_keur,
            "shl_rate": inputs.financing.shl_rate,
            "gearing_ratio": inputs.financing.gearing_ratio,
            "senior_tenor_years": inputs.financing.senior_tenor_years,
            "base_rate": inputs.financing.base_rate,
            "margin_bps": inputs.financing.margin_bps,
            "floating_share": inputs.financing.floating_share,
            "fixed_share": inputs.financing.fixed_share,
            "hedge_coverage": inputs.financing.hedge_coverage,
            "commitment_fee": inputs.financing.commitment_fee,
            "arrangement_fee": inputs.financing.arrangement_fee,
            "structuring_fee": inputs.financing.structuring_fee,
            "target_dscr": inputs.financing.target_dscr,
            "lockup_dscr": inputs.financing.lockup_dscr,
            "min_llcr": inputs.financing.min_llcr,
            "dsra_months": inputs.financing.dsra_months,
        },
        "tax": {
            "corporate_rate": inputs.tax.corporate_rate,
            "loss_carryforward_years": inputs.tax.loss_carryforward_years,
            "loss_carryforward_cap": inputs.tax.loss_carryforward_cap,
            "legal_reserve_cap": inputs.tax.legal_reserve_cap,
            "thin_cap_enabled": inputs.tax.thin_cap_enabled,
            "thin_cap_de_ratio": inputs.tax.thin_cap_de_ratio,
            "atad_ebitda_limit": inputs.tax.atad_ebitda_limit,
            "atad_min_interest_keur": inputs.tax.atad_min_interest_keur,
            "wht_sponsor_dividends": inputs.tax.wht_sponsor_dividends,
            "wht_sponsor_shl_interest": inputs.tax.wht_sponsor_shl_interest,
            "shl_cap_applies": inputs.tax.shl_cap_applies,
        },
        "capex": {
            "idc_keur": inputs.capex.idc_keur,
            "commitment_fees_keur": inputs.capex.commitment_fees_keur,
            "bank_fees_keur": inputs.capex.bank_fees_keur,
            "other_financial_keur": inputs.capex.other_financial_keur,
            "vat_costs_keur": inputs.capex.vat_costs_keur,
            "reserve_accounts_keur": inputs.capex.reserve_accounts_keur,
        },
    }
    
    with open(path, "w") as f:
        json.dump(data, f, indent=2)


def load_inputs_from_json(path: str) -> "ProjectInputs":
    """Load ProjectInputs from JSON file.
    
    Args:
        path: JSON file path
    
    Returns:
        ProjectInputs instance
    """
    with open(path, "r") as f:
        data = json.load(f)
    
    # Reconstruct dates
    data["info"]["financial_close"] = date.fromisoformat(data["info"]["financial_close"])
    data["info"]["cod_date"] = date.fromisoformat(data["info"]["cod_date"])
    
    return create_project_inputs_from_excel_dict(data)